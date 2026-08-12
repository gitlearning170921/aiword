# -*- coding: utf-8 -*-
"""发补记录 Excel 导入/模板（页面0）。"""
from __future__ import annotations

import io
import re
from datetime import date, datetime
from typing import Any, Optional

from openpyxl import Workbook, load_workbook

# 表头（中文）→ 内部字段
_HEADER_MAP: dict[str, str] = {
    "所属项目": "project_name",
    "项目名称": "project_name",
    "项目": "project_name",
    "project": "project_name",
    "project name": "project_name",
    "发补意见": "opinion_text",
    "意见": "opinion_text",
    "opinion": "opinion_text",
    "优先级": "priority",
    "priority": "priority",
    "整改方案": "remediation_plan",
    "方案": "remediation_plan",
    "plan": "remediation_plan",
    "发补时间": "issued_on",
    "发补日期": "issued_on",
    "issued_on": "issued_on",
    "issued on": "issued_on",
    "整改状态": "remediation_status",
    "状态": "remediation_status",
    "status": "remediation_status",
    "整改完成时间": "completed_on",
    "整改完成日期": "completed_on",
    "完成时间": "completed_on",
    "完成日期": "completed_on",
    "completed_on": "completed_on",
    "发补类型": "deficiency_type",
    "类型": "deficiency_type",
    "type": "deficiency_type",
    "发补来源": "deficiency_source",
    "来源": "deficiency_source",
    "source": "deficiency_source",
    "注册国家": "registration_country",
    "注册国": "registration_country",
    "国家": "registration_country",
    "registration_country": "registration_country",
    "registered country": "registration_country",
    "注册类别": "registration_category",
    "类别": "registration_category",
    "registration_category": "registration_category",
    "registered category": "registration_category",
}

_TEMPLATE_HEADERS = [
    "所属项目",
    "发补意见",
    "优先级",
    "整改方案",
    "发补日期",
    "整改状态",
    "整改完成日期",
    "发补类型",
    "发补来源",
    "注册国家",
    "注册类别",
]

_TEMPLATE_EXAMPLE = [
    "示例血糖仪二类项目",
    "请补充软件需求与风险分析的追溯关系说明。",
    "高",
    "在 SRS 与风险管理报告中补齐追溯矩阵并交叉核对编号。",
    "2024-06-12",
    "未完成",
    "",
    "注册审评发补",
    "器审中心",
    "中国",
    "第二类",
]


def _norm_header(h: Any) -> str:
    s = str(h or "").strip().lower().replace(" ", "")
    s = s.replace("（", "(").replace("）", ")")
    return s


def _lookup_field(header: Any) -> Optional[str]:
    raw = str(header or "").strip()
    if not raw:
        return None
    if raw in _HEADER_MAP:
        return _HEADER_MAP[raw]
    compact = _norm_header(raw)
    for k, v in _HEADER_MAP.items():
        if _norm_header(k) == compact:
            return v
    return None


def build_deficiency_import_template_bytes() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "发补记录"
    ws.append(_TEMPLATE_HEADERS)
    ws.append(_TEMPLATE_EXAMPLE)
    ws2 = wb.create_sheet("填写说明")
    notes = [
        "1. 请从「发补记录」Sheet 第 2 行起填写；第 1 行为表头，勿改列名。",
        "2. 「所属项目」建议填写项目名称；若已在公司总览登记，将自动关联并带出注册国家/类别。",
        "3. 总览中尚无该项目时仍可导入：将按「所属项目」原文归档；建议同时填写「注册国家」「注册类别」，否则下游暂无法按维度注入。",
        "4. 若既匹配到总览项目、Excel 又填写了注册国家/类别：以 Excel 为准（可覆盖项目带出值）。",
        "5. 若与系统已有记录重复：导入时可选择「覆盖更新」或「新增重复」；手工新增时同样可选。",
        "6. 列表默认按 Excel 行序展示（与文控台账导入顺序一致）；手工新增记录排在导入记录之后。",
        "5. 优先级：高 / 中 / 低（或 high / medium / low）。",
        "6. 整改状态：未完成 / 已完成（或 open / done）。已完成时建议填写整改完成日期，缺省为导入当天。",
        "7. 发补类型：注册审评发补 / 受理发补 / 体考发补（受理/审评均按注册审评类入库）。",
        "8. 日期格式：优先 YYYY-MM-DD 或 Excel 日期；若只填年份（如 2025）将按该年 12 月 31 日导入。",
        "9. 优先级数字：1=高，2=中，3=低。",
        "10. 可增加「序号」列（导入时忽略）。示例行可删除后导入。",
    ]
    for line in notes:
        ws2.append([line])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _parse_date_cell(val: Any) -> Optional[str]:
    """解析日期：支持 datetime/date、YYYY-MM-DD、YYYY/M/D、仅年份 YYYY、Excel 序列号。"""
    if val is None or val == "":
        return None
    if isinstance(val, datetime):
        return val.date().isoformat()
    if isinstance(val, date):
        return val.isoformat()
    # Excel 有时把日期存成 float 序列号（data_only 时）
    if isinstance(val, int):
        # 仅年份：2025 → 2025-12-31
        if 1900 <= val <= 2100:
            return date(val, 12, 31).isoformat()
        # Excel 序列号整数
        if 1 <= val <= 100000:
            try:
                from openpyxl.utils.datetime import from_excel

                return from_excel(val).date().isoformat()
            except Exception:
                return None
        return None
    if isinstance(val, float):
        # 纯年份浮点（如 2025.0）优先按年末处理，避免被当成 Excel 序列号
        if val == int(val) and 1900 <= int(val) <= 2100:
            return date(int(val), 12, 31).isoformat()
        try:
            from openpyxl.utils.datetime import from_excel

            return from_excel(val).date().isoformat()
        except Exception:
            return None
    s = str(val).strip()
    if not s:
        return None
    # 纯四位年份 → 该年最后一天（与业务约定：2025 表示 2025-12-31）
    if re.fullmatch(r"\d{4}", s):
        y = int(s)
        if 1900 <= y <= 2100:
            return date(y, 12, 31).isoformat()
        return None
    # 2024/6/12 or 2024-06-12
    m = re.match(r"^(\d{4})[/-](\d{1,2})[/-](\d{1,2})", s)
    if m:
        y, mo, d = int(m.group(1)), int(m.group(2)), int(m.group(3))
        try:
            return date(y, mo, d).isoformat()
        except ValueError:
            return None
    if len(s) >= 10:
        try:
            return date.fromisoformat(s[:10]).isoformat()
        except ValueError:
            return None
    return None


def _parse_priority(val: Any) -> str:
    """优先级：高/中/低；数字 1=高、2=中、3=低（常见台账习惯）。"""
    if isinstance(val, (int, float)) and val == int(val):
        n = int(val)
        if n == 1:
            return "high"
        if n == 3:
            return "low"
        return "medium"
    s = str(val or "").strip().lower()
    if s in ("高", "high", "h", "1"):
        return "high"
    if s in ("低", "low", "l", "3"):
        return "low"
    return "medium"


def _parse_status(val: Any) -> str:
    s = str(val or "").strip().lower()
    if s in ("已完成", "完成", "done", "completed", "closed", "close"):
        return "done"
    return "open"


def _parse_type(val: Any) -> str:
    s = str(val or "").strip().lower().replace(" ", "")
    if not s:
        return "registration_review"
    if "体考" in s or "型检" in s or s in ("type_testing", "typetesting", "tt"):
        return "type_testing"
    # 受理发补 / 审评发补 / 注册审评 等均归入注册审评类
    if (
        "受理" in s
        or "审评" in s
        or "注册" in s
        or s in ("registration_review", "registrationreview", "rr")
    ):
        return "registration_review"
    return "registration_review"


def parse_deficiency_excel(file_bytes: bytes) -> tuple[list[dict[str, Any]], list[str]]:
    """解析 Excel → 行字典列表；返回 (rows, warnings)。跳过空行。"""
    warnings: list[str] = []
    try:
        wb = load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
    except Exception as exc:
        raise ValueError(f"无法读取 Excel：{exc}") from exc
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration:
        raise ValueError("Excel 为空") from None
    col_index: dict[int, str] = {}
    for i, h in enumerate(header_row or []):
        field = _lookup_field(h)
        if field:
            col_index[i] = field
    if "project_name" not in col_index.values():
        raise ValueError("缺少「所属项目」列")
    if "opinion_text" not in col_index.values():
        raise ValueError("缺少「发补意见」列")
    if "issued_on" not in col_index.values():
        raise ValueError("缺少「发补日期/发补时间」列")

    out: list[dict[str, Any]] = []
    for excel_row_no, row in enumerate(rows_iter, start=2):
        if not row or all(c is None or str(c).strip() == "" for c in row):
            continue
        item: dict[str, Any] = {"_excel_row": excel_row_no}
        for i, field in col_index.items():
            if i >= len(row):
                continue
            item[field] = row[i]
        # 跳过仍为示例标题行的明显模板样例（可选：不强制）
        project_name = str(item.get("project_name") or "").strip()
        opinion = str(item.get("opinion_text") or "").strip()
        if not project_name and not opinion:
            continue
        issued = _parse_date_cell(item.get("issued_on"))
        completed = _parse_date_cell(item.get("completed_on"))
        # 仅填年份时提示
        raw_issued = item.get("issued_on")
        if issued and (
            (isinstance(raw_issued, int) and 1900 <= raw_issued <= 2100)
            or (isinstance(raw_issued, str) and re.fullmatch(r"\d{4}", str(raw_issued).strip()))
        ):
            warnings.append(
                f"第 {excel_row_no} 行：发补日期仅有年份，已按 {issued} 导入（建议改为完整 YYYY-MM-DD）"
            )
        status = _parse_status(item.get("remediation_status"))
        if status == "done" and not completed:
            completed = date.today().isoformat()
            # 完成日也是纯年份时上面已解析；若无完成日则用今天
        if status == "open":
            completed = None
        # 完成日仅年份时同样提示
        raw_completed = item.get("completed_on")
        if completed and status == "done" and (
            (isinstance(raw_completed, int) and 1900 <= raw_completed <= 2100)
            or (
                isinstance(raw_completed, str)
                and re.fullmatch(r"\d{4}", str(raw_completed).strip())
            )
        ):
            warnings.append(
                f"第 {excel_row_no} 行：整改完成日期仅有年份，已按 {completed} 导入"
            )
        parsed = {
            "_excel_row": excel_row_no,
            "project_name": project_name,
            "opinion_text": opinion,
            "priority": _parse_priority(item.get("priority")),
            "remediation_plan": str(item.get("remediation_plan") or "").strip(),
            "issued_on": issued,
            "remediation_status": status,
            "completed_on": completed,
            "deficiency_type": _parse_type(item.get("deficiency_type")),
            "deficiency_source": str(item.get("deficiency_source") or "").strip(),
            "registration_country": str(item.get("registration_country") or "").strip(),
            "registration_category": str(item.get("registration_category") or "").strip(),
        }
        if not parsed["project_name"]:
            warnings.append(f"第 {excel_row_no} 行：缺少所属项目，已跳过")
            continue
        if not parsed["opinion_text"]:
            warnings.append(f"第 {excel_row_no} 行：缺少发补意见，已跳过")
            continue
        if not parsed["issued_on"]:
            warnings.append(f"第 {excel_row_no} 行：发补日期无效，已跳过")
            continue
        out.append(parsed)
    try:
        wb.close()
    except Exception:
        pass
    if not out:
        hint = "；".join(warnings[:8]) if warnings else "请确认表头与数据行"
        raise ValueError(f"未解析到有效发补行（{hint}）")
    return out, warnings
