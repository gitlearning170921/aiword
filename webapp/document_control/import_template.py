"""文控中心 Excel 导入模板：对齐既有多 Sheet 台账格式。"""
from __future__ import annotations

import io
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

REGISTRATION_SHEET_NAME = "注册文件"

# 与 _excel_import_header_map / 历史导入表一致（表头勿改名）
TEMPLATE_HEADERS: list[str] = [
    "文件编号",
    "文件名称",
    "英文名",
    "版本号",
    "状态",
    "所属项目",
    "注册国家",
    "项目编号",
]

# Sheet 顺序与历史台账常见结构一致；「目录」仅作说明索引，导入时会跳过
_TEMPLATE_SHEET_SPECS: list[dict[str, Any]] = [
    {
        "name": "DHF",
        "note": "技术文件；同一文件编号可对应多个所属项目/注册国家/项目编号（逗号分隔）",
        "example": [
            "BPEDAPP-SRS-001",
            "软件需求规格说明书",
            "Software Requirements Specification",
            "A",
            "受控",
            "血压心电数据管理软件",
            "NMPA",
            "BPEDAPP",
        ],
    },
    {
        "name": REGISTRATION_SHEET_NAME,
        "note": "注册申报文件清单",
        "example": [
            "BPEDAPP-REG-001",
            "产品技术要求",
            "Product Technical Requirements",
            "1.0",
            "受控",
            "血压心电数据管理软件",
            "NMPA",
            "BPEDAPP",
        ],
    },
    {
        "name": "SOP",
        "note": "操作性文件",
        "example": [
            "SOP-QA-001",
            "软件发布操作规程",
            "Software Release SOP",
            "B",
            "受控",
            "血压心电数据管理软件",
            "NMPA",
            "BPEDAPP",
        ],
    },
    {
        "name": "程序文件",
        "note": "质量手册/程序/管理性文件；编号多按程序条款手工填写",
        "example": [
            "QP4.2.3",
            "文件控制程序",
            "Document Control Procedure",
            "C",
            "受控",
            "",
            "",
            "",
        ],
    },
    {
        "name": "四级表单",
        "note": "外来文件/质量记录",
        "example": [
            "QR-QP4.2.4-01",
            "受控文件清单",
            "Controlled Document List",
            "A",
            "受控",
            "",
            "",
            "",
        ],
    },
]


def _header_style_cells(ws, headers: list[str]) -> None:
    fill = PatternFill("solid", fgColor="D9E2F3")
    font = Font(bold=True)
    for col, title in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=title)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(wrap_text=True, vertical="center")
    widths = (16, 28, 28, 10, 10, 28, 12, 14)
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def build_document_control_import_template_bytes(*, include_sample: bool = True) -> bytes:
    """生成文控导入模板 xlsx。

    - 多工作表：DHF / 注册文件 / SOP / 程序文件 / 四级表单 + 填写说明
    - 可选「目录」索引 Sheet（导入逻辑会跳过）
    - include_sample=True 时每个业务 Sheet 带 1 行示例（可删后导入）
    """
    wb = Workbook()
    # 目录（导入跳过）
    ws_toc = wb.active
    ws_toc.title = "目录"
    _header_style_cells(ws_toc, ["工作表", "说明"])
    for spec in _TEMPLATE_SHEET_SPECS:
        ws_toc.append([spec["name"], spec.get("note") or ""])
    ws_toc.append(["填写说明", "请先阅读「填写说明」Sheet"])
    ws_toc.column_dimensions["A"].width = 14
    ws_toc.column_dimensions["B"].width = 72

    for spec in _TEMPLATE_SHEET_SPECS:
        ws = wb.create_sheet(spec["name"])
        _header_style_cells(ws, TEMPLATE_HEADERS)
        if include_sample and spec.get("example"):
            ws.append(list(spec["example"]))

    ws_help = wb.create_sheet("填写说明")
    notes = [
        "1. 请按工作表分类填写：DHF、注册文件、SOP、程序文件、四级表单；「目录」仅索引，导入时忽略。",
        "2. 每个业务 Sheet 第 1 行为表头，请勿修改列名；从第 2 行起填写数据。",
        "3. 必填/关键列：「文件编号」（或「编号/受控编号」）用于识别与去重；「文件名称」「版本号」建议填写。",
        "4. 「状态」：受控/有效/现行 等可导入；作废/废止/失效 等将跳过（若有状态列）。",
        "5. 「所属项目」「注册国家」「项目编号」可选；与页面1项目管理对应。同项目可多条记录共用同一项目编号。",
        "6. DHF 支持多项目作用域：同一行可用逗号分隔多个所属项目/注册国家/项目编号（按位置对齐）。",
        "7. 增量导入：已有相同文件编号时更新可识别字段；无「项目编号」列时不覆盖台账已有项目编号。",
        "8. 示例行可删除后导入；空行自动忽略。",
        "9. 文件格式：.xlsx / .xls。",
        "10. 表头同义列亦支持：编号/文件号/受控编号、名称/文档名称、version、project code 等。",
    ]
    ws_help.append(["填写说明"])
    ws_help["A1"].font = Font(bold=True)
    for line in notes:
        ws_help.append([line])
    ws_help.column_dimensions["A"].width = 100

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
