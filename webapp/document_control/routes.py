from __future__ import annotations

import io
import json
import re
import time
import uuid
from datetime import date, datetime
from typing import Any, Iterable, Optional

import requests
from flask import Blueprint, current_app, jsonify, render_template, request, session
from openpyxl import load_workbook
from sqlalchemy import case, func, select
from sqlalchemy.exc import IntegrityError

from webapp import db
from webapp._integration_common import (
    format_upstream_request_error,
    integration_api_base,
    integration_request,
    integration_requests_timeout,
    login_wall,
    upstream_headers,
)
from webapp.app_settings import is_effective_feature_enabled
from webapp.models import (
    ControlledDocument,
    DocumentControlImportLog,
    GenerationSummary,
    NumberAllocation,
    NumberingScheme,
    Project,
    UploadRecord,
    VersionTaskGenerationJob,
    now_local,
)
from webapp.tenant_context import resolve_organization_context
from webapp.user_facing import api_debug_fields, user_facing_upstream_error
from webapp.authz import is_page13_super_admin
from .version_task_generator import (
    build_adjustment_rows,
    delete_project_version_record,
    diagnose_release_dates,
    feedback_rows_for_org,
    generate_task_preview,
    get_latest_version_task_preview,
    get_project_version_record,
    list_project_version_records,
    batch_save_project_version_records,
    rebind_project_version_records,
    resolve_project_product_name,
    save_project_version_record_item,
    save_version_task_preview_edits,
    set_project_product_name,
    suggest_release_dates,
    upsert_project_version_records,
)
from .allocation_categories import (
    enrich_issue_categories,
    resolve_scheme_for_issue,
    sheet_category_for_doc_type,
)
from .kb_rules import upsert_schemes_from_kb_rules
from .numbering_engine import (
    controlled_number_conflict_message,
    issue_number,
    load_controlled_docs_by_norm,
    normalize_document_number,
    preview_next_number,
    registration_compare_key,
    release_expired_number_reservations,
    reserve_number,
    scheme_allocation_prefix,
    _scheme_uses_subtype_template,
)
from .subtype_resolver import SubtypeChoiceRequired, find_existing_controlled_doc_by_title
from .title_en_resolver import persist_title_en_cache, resolve_title_en_for_issue
from webapp.project_code_uniqueness import gate_project_code_save
from .project_link import (
    backfill_controlled_document_project_ids,
    link_controlled_document_to_page1_project,
)


document_control_bp = Blueprint("document_control", __name__)

_DOC_STATUS_CONTROLLED = "controlled"
_DOC_STATUS_VOIDED = "voided"
_REGISTRATION_SHEET_NAME = "注册文件"
_SCOPE_DISPLAY_SEP = ", "
# 跨 Sheet 增量更新时，以下分类的名称/英文名不被其他 Sheet 覆盖
_AUTHORITATIVE_NAME_SHEETS = frozenset({"程序文件", "四级表单"})


def _parse_registration_submitted_filter(raw: Optional[str]) -> Optional[bool]:
    value = (raw or "").strip().lower()
    if value in ("1", "true", "yes", "on", "已递交"):
        return True
    if value in ("0", "false", "no", "off", "未递交"):
        return False
    return None


def _parse_registration_submitted_value(raw: Any) -> bool:
    if isinstance(raw, bool):
        return raw
    parsed = _parse_registration_submitted_filter(
        str(raw).strip() if raw is not None else ""
    )
    return bool(parsed) if parsed is not None else False


def _parse_status_filter(raw: Optional[str]) -> Optional[str]:
    value = (raw or "").strip().lower()
    if value in (_DOC_STATUS_VOIDED, "obsolete", "作废"):
        return _DOC_STATUS_VOIDED
    if value in (_DOC_STATUS_CONTROLLED, "active", "受控"):
        return _DOC_STATUS_CONTROLLED
    return None


def _project_code_gate_response(
    org_id: Optional[str],
    project_code: Optional[str],
    *,
    project_id: Optional[str] = None,
    project_name: Optional[str] = None,
    registered_country: Optional[str] = None,
    exclude_upload_id: Optional[str] = None,
    exclude_document_id: Optional[str] = None,
    confirm_sync: bool = False,
):
    code = (project_code or "").strip()
    if not code:
        return None
    gate = gate_project_code_save(
        org_id,
        code,
        project_id=project_id,
        project_name=project_name,
        registered_country=registered_country,
        exclude_upload_id=exclude_upload_id,
        exclude_document_id=exclude_document_id,
        confirm_sync=confirm_sync,
    )
    if not gate:
        return None
    if gate[0] == "confirm":
        return jsonify(gate[1]), 409
    return jsonify({"message": gate[1]}), 409


def _documents_base_query(org_id: str):
    return ControlledDocument.query.filter_by(organization_id=org_id)


def _excel_import_batch_sort_at_subquery():
    """文档所属 Excel 导入批次的最早日志时间（同批次共享，用于批次间排序）。"""
    return (
        select(func.min(DocumentControlImportLog.created_at))
        .where(
            DocumentControlImportLog.import_batch_id == ControlledDocument.import_batch_id,
            DocumentControlImportLog.import_batch_id.isnot(None),
            DocumentControlImportLog.import_batch_id != "",
        )
        .correlate(ControlledDocument)
        .scalar_subquery()
    )


def _document_list_order_clauses(sheet_category: str = ""):
    """台账列表排序。

    - Excel 导入：先按导入批次时间升序（后导入批次在后），同批次按 excel 行号
    - 注册文件视图：注册 Sheet 自身记录用 registration_excel_row_index
    - 手动/申请编号等无行序记录：排在当前分类最后，按 created_at 升序
    """
    manual_tier = case(
        (ControlledDocument.excel_row_index.is_(None), 1),
        else_=0,
    )
    batch_sort_at = func.coalesce(
        _excel_import_batch_sort_at_subquery(),
        ControlledDocument.created_at,
    )
    tail = (
        ControlledDocument.created_at.asc(),
        ControlledDocument.document_number.asc(),
    )
    category = (sheet_category or "").strip()
    if category == _REGISTRATION_SHEET_NAME:
        sort_key = case(
            (
                ControlledDocument.sheet_category == _REGISTRATION_SHEET_NAME,
                ControlledDocument.registration_excel_row_index,
            ),
            else_=ControlledDocument.excel_row_index,
        )
        return (
            manual_tier.asc(),
            batch_sort_at.asc(),
            sort_key.is_(None),
            sort_key.asc(),
            *tail,
        )
    return (
        ControlledDocument.sheet_category.asc(),
        manual_tier.asc(),
        batch_sort_at.asc(),
        ControlledDocument.excel_row_index.is_(None),
        ControlledDocument.excel_row_index.asc(),
        *tail,
    )


def _excel_row_index_from_row(row: dict[str, Any]) -> Optional[int]:
    value = int(row.get("rowIndex") or 0)
    return value if value > 0 else None


def _split_scope_display_tokens(value: str) -> list[str]:
    text = (value or "").strip()
    if not text:
        return []
    parts = re.split(r"[,，、]|\s*/\s*", text)
    return [p.strip() for p in parts if p and p.strip()]


def _unique_scope_tokens(tokens: Iterable[str]) -> list[str]:
    out: list[str] = []
    seen: set[str] = set()
    for raw in tokens:
        token = (raw or "").strip()
        if not token:
            continue
        key = token.casefold()
        if key in seen:
            continue
        seen.add(key)
        out.append(token)
    return out


def _scope_display_tokens(value: str) -> list[str]:
    return _unique_scope_tokens(_split_scope_display_tokens(value))


def _format_scope_display(value: str) -> str:
    return _SCOPE_DISPLAY_SEP.join(_scope_display_tokens(value))


def _normalize_scope_field_value(value: Any) -> Optional[str]:
    text = _format_scope_display(str(value or ""))
    return text or None


def _expand_scope_pairs(
    project_name: str,
    registered_country: str,
    project_code: str = "",
) -> list[tuple[str, str, str]]:
    pn_tokens = _scope_display_tokens(project_name)
    rc_tokens = _scope_display_tokens(registered_country)
    pc_tokens = _scope_display_tokens(project_code)
    if not pn_tokens and not rc_tokens and not pc_tokens:
        return []

    if len(pn_tokens) > 1 and len(rc_tokens) > 1:
        count = max(len(pn_tokens), len(rc_tokens))
    elif len(pn_tokens) > 1:
        count = len(pn_tokens)
    elif len(rc_tokens) > 1:
        count = len(rc_tokens)
    elif len(pc_tokens) > 1:
        count = len(pc_tokens)
    else:
        count = 1

    def _token_at(tokens: list[str], index: int) -> str:
        if not tokens:
            return ""
        if len(tokens) == 1:
            return tokens[0]
        return tokens[index] if index < len(tokens) else ""

    pairs: list[tuple[str, str, str]] = []
    for i in range(count):
        pairs.append(
            (
                _token_at(pn_tokens, i),
                _token_at(rc_tokens, i),
                _token_at(pc_tokens, i),
            )
        )
    return pairs


def _registration_scope_entries(meta: Optional[dict]) -> list[dict[str, Any]]:
    if not isinstance(meta, dict):
        return []
    items = meta.get("registrationProjects")
    if not isinstance(items, list):
        return []
    return [x for x in items if isinstance(x, dict)]


def _scope_pair_key(project_name: str, registered_country: str) -> tuple[str, str]:
    return ((project_name or "").strip(), (registered_country or "").strip())


def _add_registration_scope(
    meta: dict[str, Any],
    *,
    project_name: str,
    registered_country: str,
    row_index: Optional[int],
    project_code: str = "",
) -> dict[str, Any]:
    pairs = _expand_scope_pairs(project_name, registered_country, project_code)
    if not pairs:
        return meta
    entries = _registration_scope_entries(meta)
    for pn, rc, pc in pairs:
        pair = _scope_pair_key(pn, rc)
        if pair == ("", "") and not pc:
            continue
        matched = False
        for entry in entries:
            if _scope_pair_key(entry.get("projectName", ""), entry.get("registeredCountry", "")) == pair:
                matched = True
                if pc and not (entry.get("projectCode") or "").strip():
                    entry["projectCode"] = pc
                break
        if not matched:
            entries.append(
                {
                    "projectName": pn,
                    "registeredCountry": rc,
                    "projectCode": pc,
                    "rowIndex": row_index,
                }
            )
    entries.sort(
        key=lambda e: (
            int(e.get("rowIndex") or 999_999),
            (e.get("projectName") or ""),
            (e.get("registeredCountry") or ""),
        )
    )
    meta["registrationProjects"] = entries
    return meta


def _capture_primary_scope_to_meta(doc: ControlledDocument, meta: dict[str, Any]) -> dict[str, Any]:
    if _registration_scope_entries(meta):
        return meta
    pn = (doc.project_name or "").strip()
    rc = (doc.registered_country or "").strip()
    pc = (doc.project_code or "").strip()
    if not pn and not rc and not pc:
        return meta
    return _add_registration_scope(
        meta,
        project_name=pn,
        registered_country=rc,
        row_index=doc.excel_row_index,
        project_code=pc,
    )


def _apply_registration_scope_display(doc: ControlledDocument, meta: dict[str, Any]) -> None:
    entries = _registration_scope_entries(meta)
    if not entries:
        return
    projects: list[str] = []
    countries: list[str] = []
    project_codes: list[str] = []
    seen_p: set[str] = set()
    seen_c: set[str] = set()
    seen_pc: set[str] = set()
    for entry in entries:
        for pn, rc, pc in _expand_scope_pairs(
            entry.get("projectName", "") or "",
            entry.get("registeredCountry", "") or "",
            entry.get("projectCode", "") or "",
        ):
            if pn:
                key = pn.casefold()
                if key not in seen_p:
                    projects.append(pn)
                    seen_p.add(key)
            if rc:
                key = rc.casefold()
                if key not in seen_c:
                    countries.append(rc)
                    seen_c.add(key)
            if pc:
                key = pc.casefold()
                if key not in seen_pc:
                    project_codes.append(pc)
                    seen_pc.add(key)
    if projects:
        doc.project_name = _SCOPE_DISPLAY_SEP.join(projects)
    if countries:
        doc.registered_country = _SCOPE_DISPLAY_SEP.join(countries)
    if project_codes:
        doc.project_code = _SCOPE_DISPLAY_SEP.join(project_codes)


def _registration_projects_for_doc(
    doc: ControlledDocument, meta: Optional[dict]
) -> list[dict[str, Any]]:
    entries = _registration_scope_entries(meta)
    if entries:
        items: list[dict[str, Any]] = []
        seen_pairs: set[tuple[tuple[str, str], str]] = set()
        for e in entries:
            for pn, rc, pc in _expand_scope_pairs(
                e.get("projectName", "") or "",
                e.get("registeredCountry", "") or "",
                e.get("projectCode", "") or "",
            ):
                pair = _scope_pair_key(pn, rc)
                if pair == ("", "") and not pc:
                    continue
                dedupe_key = (pair, pc)
                if dedupe_key in seen_pairs:
                    continue
                seen_pairs.add(dedupe_key)
                items.append(
                    {
                        "projectName": pn,
                        "registeredCountry": rc,
                        "projectCode": pc,
                        "rowIndex": e.get("rowIndex"),
                    }
                )
        return items
    pn = (doc.project_name or "").strip()
    rc = (doc.registered_country or "").strip()
    pc = (doc.project_code or "").strip()
    pairs = _expand_scope_pairs(pn, rc, pc)
    if not pairs:
        return []
    return [
        {
            "projectName": pair_pn,
            "registeredCountry": pair_rc,
            "projectCode": pair_pc,
            "rowIndex": doc.excel_row_index,
        }
        for pair_pn, pair_rc, pair_pc in pairs
    ]


def _scope_display_fields_from_meta(
    meta: dict[str, Any],
) -> tuple[Optional[str], Optional[str], Optional[str]]:
    scratch = ControlledDocument(
        document_number="-",
        normalized_document_number="-",
        title="-",
    )
    _apply_registration_scope_display(scratch, meta)
    return scratch.project_name, scratch.registered_country, scratch.project_code


def _set_registration_excel_row_index(doc: ControlledDocument, row_index: Optional[int]) -> None:
    """仅注册文件 Sheet 自身记录维护注册行序；不改动 DHF 等原 Sheet 的 excel_row_index。"""
    if (doc.sheet_category or "").strip() != _REGISTRATION_SHEET_NAME:
        return
    if row_index is None:
        return
    current = doc.registration_excel_row_index
    if current is None or row_index < current:
        doc.registration_excel_row_index = row_index


def _apply_document_filters(q):
    keyword = (request.args.get("keyword") or "").strip()
    project_code = (request.args.get("projectCode") or "").strip()
    project_name = (request.args.get("projectName") or "").strip()
    registered_country = (request.args.get("registeredCountry") or "").strip()
    status = _parse_status_filter(request.args.get("status"))
    registration_submitted = _parse_registration_submitted_filter(
        request.args.get("registrationSubmitted")
    )
    if keyword:
        like = f"%{keyword}%"
        q = q.filter(
            db.or_(
                ControlledDocument.document_number.ilike(like),
                ControlledDocument.title.ilike(like),
                ControlledDocument.title_en.ilike(like),
                ControlledDocument.project_code.ilike(like),
                ControlledDocument.project_name.ilike(like),
                ControlledDocument.registered_country.ilike(like),
                ControlledDocument.sheet_category.ilike(like),
            )
        )
    if project_code:
        q = q.filter(ControlledDocument.project_code == project_code)
    if project_name:
        q = q.filter(ControlledDocument.project_name.ilike(f"%{project_name}%"))
    if registered_country:
        q = q.filter(ControlledDocument.registered_country.ilike(f"%{registered_country}%"))
    if status:
        q = q.filter(ControlledDocument.status == status)
    if registration_submitted is True:
        q = q.filter(ControlledDocument.registration_submitted.is_(True))
    elif registration_submitted is False:
        q = q.filter(ControlledDocument.registration_submitted.is_(False))
    return q


def _apply_sheet_category_filter(q, sheet_category: str):
    category = (sheet_category or "").strip()
    if not category:
        return q
    if category == _REGISTRATION_SHEET_NAME:
        return q.filter(
            db.or_(
                ControlledDocument.sheet_category == _REGISTRATION_SHEET_NAME,
                ControlledDocument.registration_submitted.is_(True),
            )
        )
    return q.filter(ControlledDocument.sheet_category == category)


def _filtered_category_counts(org_id: str) -> dict[str, int]:
    """在当前筛选条件下按分类统计条数（注册文件单独规则，其余一次 GROUP BY）。"""
    base = _apply_document_filters(_documents_base_query(org_id))
    rows = (
        base.filter(
            ControlledDocument.sheet_category.isnot(None),
            ControlledDocument.sheet_category != "",
        )
        .with_entities(
            ControlledDocument.sheet_category.label("cat"),
            func.count(ControlledDocument.id).label("cnt"),
        )
        .group_by(ControlledDocument.sheet_category)
        .all()
    )
    counts: dict[str, int] = {
        (row.cat or "").strip(): int(row.cnt or 0) for row in rows if (row.cat or "").strip()
    }
    reg_q = _apply_sheet_category_filter(
        _apply_document_filters(_documents_base_query(org_id)),
        _REGISTRATION_SHEET_NAME,
    )
    counts[_REGISTRATION_SHEET_NAME] = int(reg_q.count() or 0)
    return counts


def _list_sheet_categories(org_id: str) -> list[str]:
    rows = (
        db.session.query(ControlledDocument.sheet_category)
        .filter(
            ControlledDocument.organization_id == org_id,
            ControlledDocument.sheet_category.isnot(None),
            ControlledDocument.sheet_category != "",
        )
        .distinct()
        .order_by(ControlledDocument.sheet_category.asc())
        .all()
    )
    items = [r[0] for r in rows if r[0]]
    if (
        db.session.query(ControlledDocument.id)
        .filter(
            ControlledDocument.organization_id == org_id,
            ControlledDocument.registration_submitted.is_(True),
        )
        .first()
        and _REGISTRATION_SHEET_NAME not in items
    ):
        items.append(_REGISTRATION_SHEET_NAME)
    return _sort_sheet_categories(items, org_id)


def _doc_lifecycle_label(status: Optional[str]) -> str:
    s = (status or "").strip().lower()
    if s in (_DOC_STATUS_VOIDED, "obsolete", "作废"):
        return "作废"
    return "受控"


def _require_feature():
    if not is_effective_feature_enabled("FEATURE_DOCUMENT_CONTROL"):
        return jsonify({"message": "文控中心功能未开启"}), 403
    return None


def _org_context() -> tuple[str, str]:
    explicit_org = (
        request.args.get("organizationId")
        or request.form.get("organizationId")
        or (request.get_json(silent=True) or {}).get("organizationId")
        or ""
    )
    return resolve_organization_context(explicit_organization_id=explicit_org)


def _serialize_doc(row: ControlledDocument) -> dict[str, Any]:
    meta = row.metadata_json if isinstance(row.metadata_json, dict) else {}
    registration_projects = _registration_projects_for_doc(row, meta)
    return {
        "id": row.id,
        "organizationId": row.organization_id,
        "documentNumber": row.document_number,
        "normalizedDocumentNumber": row.normalized_document_number,
        "version": row.version,
        "title": row.title,
        "titleEn": row.title_en or meta.get("titleEn"),
        "docTypeCode": row.doc_type_code,
        "projectId": row.project_id,
        "projectCode": row.project_code,
        "projectName": _format_scope_display(row.project_name or ""),
        "registeredCountry": _format_scope_display(row.registered_country or ""),
        "registrationProjects": registration_projects,
        "hasMultipleRegistrationScopes": len(registration_projects) > 1,
        "sheetCategory": row.sheet_category,
        "registrationSubmitted": bool(row.registration_submitted),
        "status": row.status,
        "statusLabel": _doc_lifecycle_label(row.status),
        "source": row.source,
        "extractConfidence": row.extract_confidence,
        "uploadRecordId": row.upload_record_id,
        "importBatchId": row.import_batch_id,
        "createdAt": row.created_at.isoformat() if row.created_at else None,
    }


def _serialize_scheme(row: NumberingScheme) -> dict[str, Any]:
    tmpl = (row.render_template or "").strip()
    auto = bool(row.is_active and tmpl and "{seq" in tmpl)
    return {
        "id": row.id,
        "name": row.name,
        "docTypeCode": row.doc_type_code,
        "patternRegex": row.pattern_regex,
        "renderTemplate": row.render_template,
        "prefixSource": row.prefix_source,
        "fixedPrefix": row.fixed_prefix,
        "seqScope": row.seq_scope,
        "seqStart": row.seq_start,
        "seqPad": row.seq_pad,
        "isActive": bool(row.is_active),
        "kbRuleExcerpt": row.kb_rule_excerpt,
        "sheetCategory": sheet_category_for_doc_type(row.doc_type_code or ""),
        "autoAllocatable": auto,
        "needsProjectCode": (row.prefix_source or "") == "from_project_code",
        "needsSubtype": "{subtype}" in tmpl,
    }


@document_control_bp.get("/document-control")
def document_control_page():
    blocked = _require_feature()
    if blocked is not None:
        return blocked
    wall = login_wall()
    if wall is not None:
        return wall
    return render_template("document_control.html")


@document_control_bp.get("/document-control/import-logs")
def document_control_import_logs_page():
    blocked = _require_feature()
    if blocked is not None:
        return blocked
    wall = login_wall()
    if wall is not None:
        return wall
    return render_template("document_control_import_logs.html")


@document_control_bp.get("/document-control/version-task-generator")
def version_task_generator_page():
    blocked = _require_feature()
    if blocked is not None:
        return blocked
    wall = login_wall()
    if wall is not None:
        return wall
    return render_template("version_task_generator.html")


def _parse_optional_date(value: Any) -> Optional[date]:
    text = (str(value or "")).strip()
    if not text:
        return None
    try:
        return datetime.strptime(text, "%Y-%m-%d").date()
    except ValueError:
        return None


def _ensure_generation_summary(upload: UploadRecord) -> None:
    existing = GenerationSummary.query.filter_by(upload_id=upload.id).first()
    if existing:
        return
    db.session.add(
        GenerationSummary(
            upload_id=upload.id,
            project_id=upload.project_id,
            project_name=upload.project_name or "",
            file_name=upload.file_name or "",
            author=upload.author or "",
        )
    )


def _project_in_org(project: Project, org_id: str) -> bool:
    if project.organization_id and str(project.organization_id).strip() != org_id:
        return False
    return True


def _suggest_release_dates_upstream(
  *,
  org_id: str,
  collection: str,
  payload: dict[str, Any],
) -> tuple[Optional[dict[str, Any]], Optional[str]]:
    base = integration_api_base()
    if not base:
        return None, user_facing_upstream_error(
            "未配置文档服务地址，无法调用发布时间检索",
            "文档服务未配置，请联系管理员",
        )
    url = f"{base}/api/integration/document-control/release-date-suggest"
    body = {
        "collection": collection,
        "productName": payload.get("productName") or "",
        "fromVersion": payload.get("fromVersion") or "",
        "toVersion": payload.get("toVersion") or "",
        "intermediateVersions": payload.get("intermediateVersions") or [],
        "targetVersion": payload.get("targetVersion"),
        "registrationCountry": payload.get("registrationCountry") or "",
    }
    current_app.logger.info("version-tasks release-date-suggest -> %s", url)
    try:
        # 直连文档服务，禁止走系统 HTTP_PROXY（否则 Clash 易 10054 后误入本地回退）
        resp = integration_request(
            "POST",
            url,
            json=body,
            headers=upstream_headers(for_multipart=False, organization_id=org_id),
            timeout=integration_requests_timeout(read_seconds=120),
        )
    except requests.RequestException as exc:
        current_app.logger.warning("version-tasks release-date-suggest upstream failed: %s", exc)
        return None, format_upstream_request_error(exc, base)
    current_app.logger.info(
        "version-tasks release-date-suggest upstream status=%s bytes=%s",
        resp.status_code,
        len(resp.content or b""),
    )
    if resp.status_code != 200:
        return None, user_facing_upstream_error(
            f"发布时间检索失败 HTTP {resp.status_code}",
            "发布时间检索失败，请稍后重试",
        )
    data = resp.json() if resp.content else {}
    if not isinstance(data, dict):
        return None, user_facing_upstream_error("上游响应非 JSON", "服务响应异常，请联系管理员")
    return data, None


def _release_date_diagnostics_summary(result: dict[str, Any]) -> dict[str, Any]:
    diag = result.get("diagnostics") if isinstance(result.get("diagnostics"), dict) else {}
    per_version = result.get("perVersion") if isinstance(result.get("perVersion"), list) else []
    return {
        "source": result.get("source") or "unknown",
        "candidateCount": len(result.get("candidates") or []),
        "versionCount": diag.get("versionCount") or len(per_version),
        "totalRawHits": diag.get("totalRawHits"),
        "versionsWithCandidates": diag.get("versionsWithCandidates"),
        "failureHint": diag.get("failureHint"),
    }


def _sanitize_release_date_diagnose(data: dict[str, Any], *, super_admin: bool) -> dict[str, Any]:
    if super_admin:
        return data
    summary = _release_date_diagnostics_summary(data)
    public_per_version = []
    for row in data.get("perVersion") or []:
        if not isinstance(row, dict):
            continue
        item_diag = row.get("diagnostics") if isinstance(row.get("diagnostics"), dict) else {}
        ddg = item_diag.get("duckduckgo") if isinstance(item_diag.get("duckduckgo"), dict) else {}
        extraction = item_diag.get("dateExtraction") if isinstance(item_diag.get("dateExtraction"), dict) else {}
        public_per_version.append(
            {
                "version": row.get("version"),
                "query": row.get("query"),
                "candidateCount": len(row.get("candidates") or []),
                "diagnostics": {
                    "rawHits": ddg.get("rawHits"),
                    "parser": ddg.get("parser"),
                    "networkError": "网络异常" if ddg.get("networkError") else None,
                    "failureHint": extraction.get("failureHint"),
                },
            }
        )
    out = {
        "mode": data.get("mode") or "diagnose",
        "fromVersion": data.get("fromVersion"),
        "toVersion": data.get("toVersion"),
        "versionChain": data.get("versionChain"),
        "candidateCount": summary.get("candidateCount"),
        "diagnostics": summary,
        "perVersion": public_per_version,
        "message": data.get("message") or summary.get("failureHint") or "诊断完成",
    }
    if data.get("route"):
        out["route"] = data.get("route")
    return out


def _diagnose_release_dates_upstream(
    *,
    org_id: str,
    collection: str,
    payload: dict[str, Any],
) -> tuple[Optional[dict[str, Any]], Optional[str], dict[str, Any]]:
    meta: dict[str, Any] = {"attempted": False, "latencyMs": 0}
    base = integration_api_base()
    meta["configured"] = bool(base)
    if not base:
        return None, user_facing_upstream_error(
            "未配置文档服务地址，无法调用发布时间检索诊断",
            "文档服务未配置，请联系管理员",
        ), meta
    url = f"{base}/api/integration/document-control/release-date-suggest/diagnose"
    body = {
        "collection": collection,
        "productName": payload.get("productName") or "",
        "fromVersion": payload.get("fromVersion") or "",
        "toVersion": payload.get("toVersion") or "",
        "intermediateVersions": payload.get("intermediateVersions") or [],
        "targetVersion": payload.get("targetVersion"),
        "registrationCountry": payload.get("registrationCountry") or "",
    }
    meta["attempted"] = True
    meta["url"] = url
    started = time.time()
    current_app.logger.info("version-tasks release-date-diagnose -> %s", url)
    try:
        resp = integration_request(
            "POST",
            url,
            json=body,
            headers=upstream_headers(for_multipart=False, organization_id=org_id),
            timeout=integration_requests_timeout(read_seconds=120),
        )
    except requests.RequestException as exc:
        meta["latencyMs"] = int((time.time() - started) * 1000)
        meta["error"] = str(exc)
        current_app.logger.warning("version-tasks release-date-diagnose upstream failed: %s", exc)
        return None, format_upstream_request_error(exc, base), meta
    meta["latencyMs"] = int((time.time() - started) * 1000)
    meta["statusCode"] = resp.status_code
    if resp.status_code != 200:
        return None, user_facing_upstream_error(
            f"发布时间检索诊断失败 HTTP {resp.status_code}",
            "发布时间检索诊断失败，请稍后重试",
        ), meta
    data = resp.json() if resp.content else {}
    if not isinstance(data, dict):
        return None, user_facing_upstream_error("上游响应非 JSON", "服务响应异常，请联系管理员"), meta
    data["route"] = "upstream"
    data["upstreamMeta"] = meta
    return data, None, meta


@document_control_bp.post("/api/document-control/version-tasks/release-date-suggest")
def api_version_task_release_date_suggest():
    blocked = _require_feature()
    if blocked is not None:
        return blocked
    wall = login_wall()
    if wall is not None:
        return wall
    org_id, collection = _org_context()
    payload = request.get_json(silent=True) or {}
    from_version = str(payload.get("fromVersion") or "").strip()
    to_version = str(payload.get("toVersion") or "").strip()
    if not from_version or not to_version:
        return jsonify({"message": "请先填写开始版本号和最新版本号"}), 400
    intermediate_versions = payload.get("intermediateVersions") or []
    if not isinstance(intermediate_versions, list):
        return jsonify({"message": "intermediateVersions 必须为数组"}), 400
    target_version = str(payload.get("targetVersion") or "").strip() or None
    registration_country = str(payload.get("registrationCountry") or "").strip()
    product_name = str(payload.get("productName") or "").strip()
    project_id = str(payload.get("projectId") or "").strip()
    if project_id:
        proj = Project.query.filter_by(id=project_id).first()
        if proj and _project_in_org(proj, org_id):
            if not registration_country:
                registration_country = str(getattr(proj, "registered_country", None) or "").strip()
            # 页面未填产品名时，回退用项目名称（通常与商店上架名接近）
            if not product_name:
                product_name = str(getattr(proj, "name", None) or "").strip()
    upstream_payload = {
        "productName": product_name,
        "fromVersion": from_version,
        "toVersion": to_version,
        "intermediateVersions": [str(x or "").strip() for x in intermediate_versions],
        "targetVersion": target_version,
        "registrationCountry": registration_country,
    }
    result, err = _suggest_release_dates_upstream(
        org_id=org_id,
        collection=collection,
        payload=upstream_payload,
    )
    if result is not None:
        result["route"] = "upstream"
        result["diagnosticsSummary"] = _release_date_diagnostics_summary(result)
        result.setdefault("productName", product_name)
        result.setdefault("registrationCountry", registration_country)
        return jsonify(result)
    # 文档服务不可达时不再做本地 DuckDuckGo（易被系统代理 10054 打断，且与手工浏览器检索不一致）
    current_app.logger.warning("version-tasks release-date-suggest upstream unavailable: %s", err)
    chain_versions: list[str] = []
    try:
        from .version_task_generator import parse_version_chain

        chain_versions = [
            x.normalized
            for x in parse_version_chain(
                from_version, to_version, upstream_payload["intermediateVersions"]
            )
        ]
    except ValueError as exc:
        return jsonify({"message": str(exc)}), 400
    if target_version:
        tv = str(target_version).strip()
        if tv:
            chain_versions = [tv]
    msg = user_facing_upstream_error(
        f"文档服务不可用，无法联网检索发布时间：{err}",
        "文档服务暂不可用，请稍后重试或手动填写各版本发布时间",
    )
    empty = {
        "ok": False,
        "source": "upstream_unreachable",
        "route": "upstream_unreachable",
        "candidates": [],
        "perVersion": [
            {
                "version": ver,
                "query": "",
                "candidates": [],
                "message": msg,
            }
            for ver in chain_versions
        ],
        "message": msg,
        "upstreamWarning": err,
        "diagnostics": {
            "failureHint": msg,
            "versionCount": len(chain_versions),
            "candidateCount": 0,
            "versionsMissing": chain_versions,
        },
    }
    empty["diagnosticsSummary"] = _release_date_diagnostics_summary(empty)
    if is_page13_super_admin():
        empty.update(api_debug_fields(upstreamWarning=err, detail=err))
    return jsonify(empty)


@document_control_bp.post("/api/document-control/version-tasks/release-date-suggest/diagnose")
def api_version_task_release_date_suggest_diagnose():
    blocked = _require_feature()
    if blocked is not None:
        return blocked
    wall = login_wall()
    if wall is not None:
        return wall
    org_id, collection = _org_context()
    super_admin = is_page13_super_admin()
    payload = request.get_json(silent=True) or {}
    from_version = str(payload.get("fromVersion") or "").strip()
    to_version = str(payload.get("toVersion") or "").strip()
    if not from_version or not to_version:
        return jsonify({"message": "请先填写开始版本号和最新版本号"}), 400
    intermediate_versions = payload.get("intermediateVersions") or []
    if not isinstance(intermediate_versions, list):
        return jsonify({"message": "intermediateVersions 必须为数组"}), 400
    target_version = str(payload.get("targetVersion") or "").strip() or None
    registration_country = str(payload.get("registrationCountry") or "").strip()
    product_name = str(payload.get("productName") or "").strip()
    project_id = str(payload.get("projectId") or "").strip()
    if project_id:
        proj = Project.query.filter_by(id=project_id).first()
        if proj and _project_in_org(proj, org_id):
            if not registration_country:
                registration_country = str(getattr(proj, "registered_country", None) or "").strip()
            if not product_name:
                product_name = str(getattr(proj, "name", None) or "").strip()
    upstream_payload = {
        "productName": product_name,
        "fromVersion": from_version,
        "toVersion": to_version,
        "intermediateVersions": [str(x or "").strip() for x in intermediate_versions],
        "targetVersion": target_version,
        "registrationCountry": registration_country,
    }

    result, err, upstream_meta = _diagnose_release_dates_upstream(
        org_id=org_id,
        collection=collection,
        payload=upstream_payload,
    )
    if result is None:
        current_app.logger.warning("version-tasks release-date-diagnose fallback local: %s", err)
        try:
            local = diagnose_release_dates(
                product_name=upstream_payload["productName"],
                from_version=from_version,
                to_version=to_version,
                intermediate_versions=upstream_payload["intermediateVersions"],
                target_version=target_version,
            )
        except ValueError as exc:
            return jsonify({"message": str(exc)}), 400
        local["route"] = "local_fallback"
        local["upstreamMeta"] = upstream_meta
        if err:
            local["upstreamWarning"] = err
        out = _sanitize_release_date_diagnose(local, super_admin=super_admin)
        if super_admin:
            out.update(api_debug_fields(upstreamWarning=err, upstreamMeta=upstream_meta))
        return jsonify(out)

    out = _sanitize_release_date_diagnose(result, super_admin=super_admin)
    if super_admin:
        out.update(api_debug_fields(upstreamMeta=upstream_meta))
    return jsonify(out)


@document_control_bp.get("/api/document-control/version-tasks/project-records")
def api_version_task_project_records_list():
    blocked = _require_feature()
    if blocked is not None:
        return blocked
    wall = login_wall()
    if wall is not None:
        return wall
    org_id, _ = _org_context()
    project_id = str(request.args.get("projectId") or "").strip()
    if not project_id:
        return jsonify({"message": "projectId 不能为空"}), 400
    project = Project.query.filter_by(id=project_id).first()
    if not project:
        return jsonify({"message": "未找到所选项目"}), 404
    if not _project_in_org(project, org_id):
        return jsonify({"message": "所选项目不在当前公司作用域内"}), 403
    items = list_project_version_records(org_id=org_id, project_id=project_id)
    product_name = resolve_project_product_name(org_id=org_id, project_id=project_id)
    return jsonify(
        {"projectId": project_id, "productName": product_name, "items": items}
    )


@document_control_bp.post("/api/document-control/version-tasks/project-product-name")
def api_version_task_project_product_name():
    """单独保存项目的应用市场产品名（刷新后回填用）。"""
    blocked = _require_feature()
    if blocked is not None:
        return blocked
    wall = login_wall()
    if wall is not None:
        return wall
    org_id, _ = _org_context()
    payload = request.get_json(silent=True) or {}
    project_id = str(payload.get("projectId") or "").strip()
    if not project_id:
        return jsonify({"message": "projectId 不能为空"}), 400
    project = Project.query.filter_by(id=project_id).first()
    if not project:
        return jsonify({"message": "未找到所选项目"}), 404
    if not _project_in_org(project, org_id):
        return jsonify({"message": "所选项目不在当前公司作用域内"}), 403
    try:
        result = set_project_product_name(
            org_id=org_id,
            project_id=project_id,
            product_name=str(payload.get("productName") or "").strip(),
        )
    except ValueError as exc:
        return jsonify({"message": str(exc)}), 400
    # 尚无版本记录/预览批次时，创建轻量草稿批次仅保存产品名，避免刷新丢失
    if int(result.get("updatedRecords") or 0) == 0 and int(result.get("updatedJobs") or 0) == 0:
        product_name = str(payload.get("productName") or "").strip()
        if product_name:
            draft = VersionTaskGenerationJob(
                organization_id=org_id,
                project_id=project_id,
                from_version="-",
                to_version="-",
                status="product_meta",
                rule_snapshot_json={"productName": product_name},
                preview_json=None,
                created_by_user_id=str(session.get("user_id") or "").strip() or None,
            )
            db.session.add(draft)
            result["updatedJobs"] = 1
            result["metaJobCreated"] = True
    db.session.commit()
    return jsonify(result)


@document_control_bp.post("/api/document-control/version-tasks/project-records")
def api_version_task_project_records_save():
    blocked = _require_feature()
    if blocked is not None:
        return blocked
    wall = login_wall()
    if wall is not None:
        return wall
    org_id, _ = _org_context()
    payload = request.get_json(silent=True) or {}
    project_id = str(payload.get("projectId") or "").strip()
    if not project_id:
        return jsonify({"message": "projectId 不能为空"}), 400
    project = Project.query.filter_by(id=project_id).first()
    if not project:
        return jsonify({"message": "未找到所选项目"}), 404
    if not _project_in_org(project, org_id):
        return jsonify({"message": "所选项目不在当前公司作用域内"}), 403
    version_release_dates = payload.get("versionReleaseDates")
    if not isinstance(version_release_dates, dict) or not version_release_dates:
        return jsonify({"message": "versionReleaseDates 不能为空"}), 400
    try:
        items = upsert_project_version_records(
            org_id=org_id,
            project_id=project_id,
            version_release_dates=version_release_dates,
            product_name=str(payload.get("productName") or "").strip(),
            chain_from_version=str(payload.get("fromVersion") or "").strip(),
            chain_to_version=str(payload.get("toVersion") or "").strip(),
            generation_status=str(payload.get("generationStatus") or "").strip().lower(),
            allow_downgrade_status=bool(payload.get("allowDowngradeStatus")),
        )
    except ValueError as exc:
        return jsonify({"message": str(exc)}), 400
    db.session.commit()
    return jsonify({"saved": len(items), "items": items})


@document_control_bp.post("/api/document-control/version-tasks/project-records/item")
def api_version_task_project_record_create():
    blocked = _require_feature()
    if blocked is not None:
        return blocked
    wall = login_wall()
    if wall is not None:
        return wall
    org_id, _ = _org_context()
    payload = request.get_json(silent=True) or {}
    project_id = str(payload.get("projectId") or "").strip()
    if not project_id:
        return jsonify({"message": "projectId 不能为空"}), 400
    project = Project.query.filter_by(id=project_id).first()
    if not project:
        return jsonify({"message": "未找到所选项目"}), 404
    if not _project_in_org(project, org_id):
        return jsonify({"message": "所选项目不在当前公司作用域内"}), 403
    version = str(payload.get("version") or "").strip()
    if not version:
        return jsonify({"message": "version 不能为空"}), 400
    try:
        item = save_project_version_record_item(
            org_id=org_id,
            project_id=project_id,
            version=version,
            released_at=str(payload.get("releasedAt") or "").strip(),
            product_name=str(payload.get("productName") or "").strip(),
            chain_from_version=str(payload.get("chainFromVersion") or payload.get("fromVersion") or "").strip(),
            chain_to_version=str(payload.get("chainToVersion") or payload.get("toVersion") or "").strip(),
            generation_status=str(payload.get("generationStatus") or "none").strip().lower(),
            allow_downgrade_status=True,
        )
    except ValueError as exc:
        return jsonify({"message": str(exc)}), 400
    db.session.commit()
    return jsonify({"item": item})


@document_control_bp.post("/api/document-control/version-tasks/project-records/batch")
def api_version_task_project_records_batch():
    blocked = _require_feature()
    if blocked is not None:
        return blocked
    wall = login_wall()
    if wall is not None:
        return wall
    org_id, _ = _org_context()
    payload = request.get_json(silent=True) or {}
    items = payload.get("items")
    if not isinstance(items, list) or not items:
        return jsonify({"message": "items 不能为空"}), 400
    project_ids = {
        str(x.get("projectId") or "").strip()
        for x in items
        if isinstance(x, dict) and str(x.get("projectId") or "").strip()
    }
    if not project_ids:
        return jsonify({"message": "items 中缺少关联项目"}), 400
    for project_id in project_ids:
        project = Project.query.filter_by(id=project_id).first()
        if not project:
            return jsonify({"message": f"未找到项目：{project_id}"}), 404
        if not _project_in_org(project, org_id):
            return jsonify({"message": "所选项目不在当前公司作用域内"}), 403
    try:
        result = batch_save_project_version_records(
            org_id=org_id,
            items=items,
            chain_from_version=str(
                payload.get("chainFromVersion") or payload.get("fromVersion") or ""
            ).strip(),
            chain_to_version=str(
                payload.get("chainToVersion") or payload.get("toVersion") or ""
            ).strip(),
        )
    except ValueError as exc:
        return jsonify({"message": str(exc)}), 400
    db.session.commit()
    return jsonify(result)


@document_control_bp.patch("/api/document-control/version-tasks/project-records/<record_id>")
def api_version_task_project_record_update(record_id: str):
    blocked = _require_feature()
    if blocked is not None:
        return blocked
    wall = login_wall()
    if wall is not None:
        return wall
    org_id, _ = _org_context()
    payload = request.get_json(silent=True) or {}
    row = get_project_version_record(org_id=org_id, record_id=record_id)
    if not row:
        return jsonify({"message": "未找到版本记录"}), 404
    target_project_id = str(payload.get("projectId") or row.project_id or "").strip()
    if not target_project_id:
        return jsonify({"message": "projectId 不能为空"}), 400
    project = Project.query.filter_by(id=target_project_id).first()
    if not project or not _project_in_org(project, org_id):
        return jsonify({"message": "所选项目不在当前公司作用域内"}), 403
    version = str(payload.get("version") or row.version).strip()
    try:
        result = rebind_project_version_records(
            org_id=org_id,
            record_id=record_id,
            target_project_id=target_project_id,
            version=version,
            released_at=str(payload.get("releasedAt") or "").strip(),
            product_name=str(payload.get("productName") or row.product_name or "").strip(),
            chain_from_version=str(payload.get("chainFromVersion") or row.chain_from_version or "").strip(),
            chain_to_version=str(payload.get("chainToVersion") or row.chain_to_version or "").strip(),
            generation_status=str(payload.get("generationStatus") or row.generation_status or "none").strip().lower(),
            allow_downgrade_status=bool(payload.get("allowDowngradeStatus", True)),
        )
    except ValueError as exc:
        return jsonify({"message": str(exc)}), 400
    db.session.commit()
    return jsonify(result)


@document_control_bp.delete("/api/document-control/version-tasks/project-records/<record_id>")
def api_version_task_project_record_delete(record_id: str):
    blocked = _require_feature()
    if blocked is not None:
        return blocked
    wall = login_wall()
    if wall is not None:
        return wall
    org_id, _ = _org_context()
    row = get_project_version_record(org_id=org_id, record_id=record_id)
    if not row:
        return jsonify({"message": "未找到版本记录"}), 404
    project = Project.query.filter_by(id=row.project_id).first()
    if not project or not _project_in_org(project, org_id):
        return jsonify({"message": "所选项目不在当前公司作用域内"}), 403
    delete_project_version_record(org_id=org_id, record_id=record_id)
    db.session.commit()
    return jsonify({"deleted": True, "id": record_id})


@document_control_bp.get("/api/document-control/version-tasks/latest-preview")
def api_version_task_latest_preview():
    blocked = _require_feature()
    if blocked is not None:
        return blocked
    wall = login_wall()
    if wall is not None:
        return wall
    org_id, _ = _org_context()
    project_id = str(request.args.get("projectId") or "").strip() or None
    if project_id:
        project = Project.query.filter_by(id=project_id).first()
        if not project:
            return jsonify({"message": "未找到所选项目"}), 404
        if not _project_in_org(project, org_id):
            return jsonify({"message": "所选项目不在当前公司作用域内"}), 403
    data = get_latest_version_task_preview(org_id=org_id, project_id=project_id)
    if not data:
        return jsonify({"jobId": "", "items": [], "message": "暂无已保存的预览"})
    return jsonify(data)


@document_control_bp.post("/api/document-control/version-tasks/preview")
def api_version_task_preview():
    blocked = _require_feature()
    if blocked is not None:
        return blocked
    wall = login_wall()
    if wall is not None:
        return wall
    org_id, _ = _org_context()
    payload = request.get_json(silent=True) or {}
    from_version = str(payload.get("fromVersion") or "").strip()
    to_version = str(payload.get("toVersion") or "").strip()
    if not from_version or not to_version:
        return jsonify({"message": "开始版本号和最新版本号不能为空"}), 400

    intermediate_versions = payload.get("intermediateVersions") or []
    if not isinstance(intermediate_versions, list):
        return jsonify({"message": "intermediateVersions 必须为数组"}), 400

    version_release_dates = payload.get("versionReleaseDates")
    if not isinstance(version_release_dates, dict) or not version_release_dates:
        # 兼容旧字段：单一 releaseDate 仅填充最新版本
        legacy_release_date = str(
            payload.get("releaseDate") or payload.get("confirmedReleaseDate") or ""
        ).strip()
        if legacy_release_date:
            version_release_dates = {to_version: legacy_release_date}
        else:
            return jsonify(
                {
                    "message": "请为版本链路中的每个版本填写发布时间后再生成预览（可先点「检索候选发布日期」填入）",
                    "needsReleaseDateConfirmation": True,
                }
            ), 409

    try:
        feedback_rows = feedback_rows_for_org(org_id)
        preview = generate_task_preview(
            from_version=from_version,
            to_version=to_version,
            intermediate_versions=[str(x or "").strip() for x in intermediate_versions],
            version_release_dates=version_release_dates,
            feedback_rows=feedback_rows,
        )
    except ValueError as exc:
        return jsonify({"message": str(exc)}), 400
    except Exception as exc:
        return jsonify({"message": f"生成预览失败：{exc}"}), 500

    project_id = str(payload.get("projectId") or "").strip() or None
    # 同一项目+起止版本再次预览：更新已有批次，避免重复堆积
    job = None
    if project_id:
        job = (
            VersionTaskGenerationJob.query.filter_by(
                organization_id=org_id,
                project_id=project_id,
                from_version=preview["fromVersion"],
                to_version=preview["toVersion"],
            )
            .order_by(VersionTaskGenerationJob.updated_at.desc())
            .first()
        )
    preview_updated = bool(job)
    if job is None:
        job = VersionTaskGenerationJob(
            organization_id=org_id,
            project_id=project_id,
            from_version=preview["fromVersion"],
            to_version=preview["toVersion"],
            created_by_user_id=str(session.get("user_id") or "").strip() or None,
        )
        db.session.add(job)
    job.project_id = project_id
    job.from_version = preview["fromVersion"]
    job.to_version = preview["toVersion"]
    job.intermediate_versions_json = (
        preview.get("versionChain")[1:-1] if preview.get("versionChain") else []
    )
    job.released_at = _parse_optional_date(preview.get("releaseDate"))
    # 已下发过的批次再次预览仍保留 applied，结果以最新 preview 为准
    if (job.status or "") != "applied":
        job.status = "previewed"
    product_name = str(payload.get("productName") or "").strip()
    job.rule_snapshot_json = {
        "rulesMode": "yy_iw_020_plus_qp739",
        "versionRule": "X.Y.Z.B",
        "ruleSource": preview.get("ruleSource"),
        "ruleBasis": preview.get("ruleBasis"),
        "versionReleaseDates": preview.get("versionReleaseDates") or {},
        "productName": product_name,
        "generatedAt": now_local().isoformat(),
        "updated": preview_updated,
    }
    preview = dict(preview)
    if product_name:
        preview["productName"] = product_name
    job.preview_json = preview
    db.session.flush()

    saved_records: list[dict[str, Any]] = []
    if project_id:
        project = Project.query.filter_by(id=project_id).first()
        if not project:
            db.session.rollback()
            return jsonify({"message": "未找到所选项目"}), 404
        if not _project_in_org(project, org_id):
            db.session.rollback()
            return jsonify({"message": "所选项目不在当前公司作用域内"}), 403
        try:
            saved_records = upsert_project_version_records(
                org_id=org_id,
                project_id=project_id,
                version_release_dates=preview.get("versionReleaseDates") or {},
                product_name=product_name,
                chain_from_version=preview.get("fromVersion") or "",
                chain_to_version=preview.get("toVersion") or "",
                generation_status="previewed",
                job_id=job.id,
            )
        except ValueError as exc:
            db.session.rollback()
            return jsonify({"message": str(exc)}), 400

    db.session.commit()
    return jsonify(
        {
            "jobId": job.id,
            "previewUpdated": preview_updated,
            "savedRecords": saved_records,
            **preview,
        }
    )


@document_control_bp.post("/api/document-control/version-tasks/feedback")
def api_version_task_feedback():
    blocked = _require_feature()
    if blocked is not None:
        return blocked
    wall = login_wall()
    if wall is not None:
        return wall
    org_id, _ = _org_context()
    payload = request.get_json(silent=True) or {}
    adjustments = payload.get("adjustments") or []
    if not isinstance(adjustments, list):
        return jsonify({"message": "adjustments 必须为数组"}), 400
    rows = build_adjustment_rows(
        org_id=org_id,
        source_job_id=str(payload.get("sourceJobId") or "").strip() or None,
        project_id=str(payload.get("projectId") or "").strip() or None,
        adjustments=adjustments,
    )
    if not rows:
        return jsonify({"saved": 0, "message": "无有效调整项"}), 200
    for row in rows:
        db.session.add(row)
    db.session.commit()
    return jsonify({"saved": len(rows)})


@document_control_bp.post("/api/document-control/version-tasks/preview/save-edits")
def api_version_task_preview_save_edits():
    """保存预览表格人工修改；同时写入反馈，供下次「生成预览」优化。"""
    blocked = _require_feature()
    if blocked is not None:
        return blocked
    wall = login_wall()
    if wall is not None:
        return wall
    org_id, _ = _org_context()
    payload = request.get_json(silent=True) or {}
    job_id = str(payload.get("jobId") or payload.get("sourceJobId") or "").strip()
    items = payload.get("items")
    if not job_id:
        return jsonify({"message": "请先生成预览后再保存修改"}), 400
    if not isinstance(items, list):
        return jsonify({"message": "items 必须为数组"}), 400
    adjustments = payload.get("adjustments")
    if adjustments is not None and not isinstance(adjustments, list):
        return jsonify({"message": "adjustments 必须为数组"}), 400
    project_id = str(payload.get("projectId") or "").strip() or None
    if project_id:
        project = Project.query.filter_by(id=project_id).first()
        if not project:
            return jsonify({"message": "未找到所选项目"}), 404
        if not _project_in_org(project, org_id):
            return jsonify({"message": "所选项目不在当前公司作用域内"}), 403
    try:
        result = save_version_task_preview_edits(
            org_id=org_id,
            job_id=job_id,
            items=items,
            adjustments=adjustments if isinstance(adjustments, list) else None,
            project_id=project_id,
        )
    except ValueError as exc:
        return jsonify({"message": str(exc)}), 400
    db.session.commit()
    return jsonify(
        {
            "message": (
                f"已保存预览修改（{result.get('itemCount') or 0} 条）"
                + (
                    f"，并写入反馈 {result.get('feedbackSaved') or 0} 条供下次生成生效"
                    if result.get("feedbackSaved")
                    else "（无字段变更，仅更新了预览快照）"
                )
            ),
            **result,
        }
    )


@document_control_bp.post("/api/document-control/version-tasks/apply")
def api_version_task_apply():
    blocked = _require_feature()
    if blocked is not None:
        return blocked
    wall = login_wall()
    if wall is not None:
        return wall
    org_id, _ = _org_context()
    payload = request.get_json(silent=True) or {}
    project_id = str(payload.get("projectId") or "").strip()
    if not project_id:
        return jsonify({"message": "转换为任务列表时必须选择项目"}), 400
    project = Project.query.filter_by(id=project_id).first()
    if not project:
        return jsonify({"message": "未找到所选项目"}), 404
    if project.organization_id and str(project.organization_id).strip() != org_id:
        return jsonify({"message": "所选项目不在当前公司作用域内"}), 403

    items = payload.get("items")
    if not isinstance(items, list) or not items:
        return jsonify({"message": "items 不能为空"}), 400

    source_job_id = str(payload.get("sourceJobId") or "").strip() or None
    source_job = (
        VersionTaskGenerationJob.query.filter_by(id=source_job_id, organization_id=org_id).first()
        if source_job_id
        else None
    )

    created = 0
    updated = 0
    upload_ids: list[str] = []
    for row in items:
        if not isinstance(row, dict):
            continue
        file_name = str(row.get("fileName") or "").strip()
        author = str(row.get("author") or "").strip()
        task_type = str(row.get("taskType") or "").strip()
        if not file_name:
            continue
        if not author:
            author = "待分配"
        if not task_type:
            task_type = "版本变更任务"
        existing = UploadRecord.query.filter_by(
            project_name=project.name,
            file_name=file_name,
            task_type=task_type,
            author=author,
        ).first()
        due_date = _parse_optional_date(row.get("dueDate"))
        document_display_date = _parse_optional_date(row.get("documentDisplayDate"))
        notes = str(row.get("notes") or "").strip() or None
        module = str(row.get("belongingModule") or "").strip() or None
        if existing:
            existing.organization_id = org_id
            existing.project_id = project.id
            existing.project_name = project.name
            existing.project_code = project.project_code
            existing.notes = notes
            existing.due_date = due_date
            existing.document_display_date = document_display_date
            existing.belonging_module = module
            existing.file_version = str(row.get("fileVersion") or "").strip() or None
            existing.registration_version = str(row.get("registrationVersion") or "").strip() or None
            existing.task_status = "pending"
            existing.completion_status = None
            db.session.add(existing)
            upload_ids.append(existing.id)
            updated += 1
        else:
            created_row = UploadRecord(
                organization_id=org_id,
                project_id=project.id,
                project_name=project.name,
                project_code=project.project_code,
                file_name=file_name,
                task_type=task_type,
                author=author,
                notes=notes,
                due_date=due_date,
                document_display_date=document_display_date,
                belonging_module=module,
                file_version=str(row.get("fileVersion") or "").strip() or None,
                registration_version=str(row.get("registrationVersion") or "").strip() or None,
                task_status="pending",
                completion_status=None,
            )
            db.session.add(created_row)
            db.session.flush()
            _ensure_generation_summary(created_row)
            upload_ids.append(created_row.id)
            created += 1

    if source_job is not None:
        source_job.project_id = project.id
        source_job.status = "applied"
        source_job.result_json = {
            "created": created,
            "updated": updated,
            "uploadIds": upload_ids,
        }
        db.session.add(source_job)

    version_release_dates = payload.get("versionReleaseDates")
    if not isinstance(version_release_dates, dict) or not version_release_dates:
        if source_job and isinstance(source_job.preview_json, dict):
            version_release_dates = source_job.preview_json.get("versionReleaseDates") or {}
    if isinstance(version_release_dates, dict) and version_release_dates:
        try:
            upsert_project_version_records(
                org_id=org_id,
                project_id=project_id,
                version_release_dates=version_release_dates,
                product_name=str(payload.get("productName") or "").strip(),
                chain_from_version=str(payload.get("fromVersion") or "").strip()
                or (source_job.from_version if source_job else ""),
                chain_to_version=str(payload.get("toVersion") or "").strip()
                or (source_job.to_version if source_job else ""),
                generation_status="generated",
                job_id=source_job_id,
            )
        except ValueError:
            pass

    db.session.commit()
    return jsonify(
        {
            "created": created,
            "updated": updated,
            "uploadIds": upload_ids,
            "message": f"已下发到任务列表：新增 {created} 条，更新 {updated} 条",
        }
    )


@document_control_bp.get("/api/document-control/bootstrap")
def api_document_control_bootstrap():
    blocked = _require_feature()
    if blocked is not None:
        return blocked
    wall = login_wall()
    if wall is not None:
        return wall
    org_id, _ = _org_context()
    schemes = (
        NumberingScheme.query.filter_by(organization_id=org_id)
        .order_by(NumberingScheme.created_at.desc())
        .all()
    )
    return jsonify(
        {
            "organizationId": org_id,
            "schemes": [_serialize_scheme(x) for x in schemes],
        }
    )


@document_control_bp.get("/api/document-control/documents")
def api_document_control_documents():
    blocked = _require_feature()
    if blocked is not None:
        return blocked
    wall = login_wall()
    if wall is not None:
        return wall
    org_id, _ = _org_context()
    try:
        backfill_controlled_document_project_ids(org_id)
        db.session.commit()
    except Exception:
        db.session.rollback()
    q = _documents_base_query(org_id)
    q = _apply_document_filters(q)
    sheet_category = (request.args.get("sheetCategory") or "").strip()
    q = _apply_sheet_category_filter(q, sheet_category)
    page = max(1, int(request.args.get("page") or 1))
    page_size = min(200, max(10, int(request.args.get("pageSize") or 50)))
    total = q.count()
    rows = (
        q.order_by(*_document_list_order_clauses(sheet_category))
        .offset((page - 1) * page_size)
        .limit(page_size)
        .all()
    )
    return jsonify(
        {
            "items": [_serialize_doc(x) for x in rows],
            "total": total,
            "page": page,
            "pageSize": page_size,
        }
    )


def _get_org_document(doc_id: str, org_id: str) -> Optional[ControlledDocument]:
    return ControlledDocument.query.filter_by(id=doc_id, organization_id=org_id).first()


def _apply_document_payload(doc: ControlledDocument, data: dict[str, Any]) -> Optional[str]:
    if "title" in data:
        title = (data.get("title") or "").strip()
        if not title:
            return "请填写文件名称"
        doc.title = title
    if "titleEn" in data:
        doc.title_en = (data.get("titleEn") or "").strip() or None
    if "version" in data:
        doc.version = (data.get("version") or "").strip() or None
    if "docTypeCode" in data:
        doc.doc_type_code = (data.get("docTypeCode") or "").strip() or None
    if "projectCode" in data:
        doc.project_code = (data.get("projectCode") or "").strip() or None
    if "projectName" in data:
        doc.project_name = _normalize_scope_field_value(data.get("projectName"))
    if "registeredCountry" in data:
        doc.registered_country = _normalize_scope_field_value(data.get("registeredCountry"))
    if "sheetCategory" in data:
        doc.sheet_category = (data.get("sheetCategory") or "").strip() or None
    if "registrationSubmitted" in data:
        doc.registration_submitted = bool(data.get("registrationSubmitted"))
    if "status" in data:
        status = _parse_status_filter(str(data.get("status") or ""))
        if status:
            doc.status = status
    meta = dict(doc.metadata_json or {})
    if doc.title_en:
        meta["titleEn"] = doc.title_en
    elif "titleEn" in data:
        meta.pop("titleEn", None)
    doc.metadata_json = meta or None
    doc.updated_at = now_local()
    return None


@document_control_bp.post("/api/document-control/documents")
def api_document_control_create_document():
    blocked = _require_feature()
    if blocked is not None:
        return blocked
    wall = login_wall()
    if wall is not None:
        return wall
    org_id, _ = _org_context()
    data = request.get_json(silent=True) or {}
    doc_num = (data.get("documentNumber") or "").strip()
    norm = normalize_document_number(doc_num)
    if not norm:
        return jsonify({"message": "请填写有效文件编号"}), 400
    title = (data.get("title") or "").strip()
    if not title:
        return jsonify({"message": "请填写文件名称"}), 400
    doc_status = _parse_status_filter(str(data.get("status") or "")) or _DOC_STATUS_CONTROLLED
    if doc_status == _DOC_STATUS_CONTROLLED:
        conflict_msg = controlled_number_conflict_message(org_id, norm)
        if conflict_msg:
            return jsonify({"message": conflict_msg}), 409
    pc = (data.get("projectCode") or "").strip() or None
    if pc:
        blocked = _project_code_gate_response(
            org_id,
            pc,
            project_id=(data.get("projectId") or "").strip() or None,
            project_name=(data.get("projectName") or "").strip() or None,
            registered_country=(data.get("registeredCountry") or "").strip() or None,
            confirm_sync=bool(data.get("confirmProjectCodeSync")),
        )
        if blocked:
            return blocked
    title_en = (data.get("titleEn") or "").strip()
    metadata_json = {"titleEn": title_en} if title_en else None
    doc = ControlledDocument(
        organization_id=org_id,
        document_number=doc_num,
        normalized_document_number=norm,
        version=(data.get("version") or "").strip() or None,
        title=title,
        title_en=title_en or None,
        doc_type_code=(data.get("docTypeCode") or "").strip() or None,
        project_id=(data.get("projectId") or "").strip() or None,
        project_code=(data.get("projectCode") or "").strip() or None,
        project_name=(data.get("projectName") or "").strip() or None,
        registered_country=(data.get("registeredCountry") or "").strip() or None,
        sheet_category=(data.get("sheetCategory") or "").strip() or None,
        registration_submitted=_parse_registration_submitted_value(data.get("registrationSubmitted")),
        status=doc_status,
        source="manual",
        metadata_json=metadata_json,
        created_by_user_id=(session.get("user_id") or "").strip() or None,
    )
    try:
        db.session.add(doc)
        db.session.flush()
        link_controlled_document_to_page1_project(doc, organization_id=org_id)
        db.session.commit()
    except IntegrityError:
        db.session.rollback()
        return jsonify({"message": "保存失败：受控编号冲突"}), 409
    return jsonify({"message": "已新增", "item": _serialize_doc(doc)})


def _controlled_document_from_manual_item(
    org_id: str,
    data: dict[str, Any],
    *,
    user_id: Optional[str],
    existing_norms: Optional[set[str]] = None,
    confirm_sync: bool = False,
) -> tuple[Optional[ControlledDocument], Optional[str]]:
    doc_num = (data.get("documentNumber") or "").strip()
    norm = normalize_document_number(doc_num)
    if not norm:
        return None, "请填写有效文件编号"
    title = (data.get("title") or "").strip()
    if not title:
        return None, "请填写文件名称"
    if existing_norms is not None and norm in existing_norms:
        return None, "本批中文件编号重复"
    doc_status = _parse_status_filter(str(data.get("status") or "")) or _DOC_STATUS_CONTROLLED
    if doc_status == _DOC_STATUS_CONTROLLED:
        conflict_msg = controlled_number_conflict_message(org_id, norm)
        if conflict_msg:
            return None, conflict_msg
    pc = (data.get("projectCode") or "").strip() or None
    if pc:
        gate = gate_project_code_save(
            org_id,
            pc,
            project_id=(data.get("projectId") or "").strip() or None,
            project_name=(data.get("projectName") or "").strip() or None,
            registered_country=(data.get("registeredCountry") or "").strip() or None,
            confirm_sync=confirm_sync,
        )
        if gate:
            if gate[0] == "confirm":
                return None, gate[1].get("message") or "项目编号变更需确认"
            return None, gate[1]
    title_en = (data.get("titleEn") or "").strip()
    metadata_json = {"titleEn": title_en} if title_en else None
    doc = ControlledDocument(
        organization_id=org_id,
        document_number=doc_num,
        normalized_document_number=norm,
        version=(data.get("version") or "").strip() or None,
        title=title,
        title_en=title_en or None,
        doc_type_code=(data.get("docTypeCode") or "").strip() or None,
        project_id=(data.get("projectId") or "").strip() or None,
        project_code=(data.get("projectCode") or "").strip() or None,
        project_name=(data.get("projectName") or "").strip() or None,
        registered_country=(data.get("registeredCountry") or "").strip() or None,
        sheet_category=(data.get("sheetCategory") or "").strip() or None,
        registration_submitted=_parse_registration_submitted_value(data.get("registrationSubmitted")),
        status=doc_status,
        source="manual",
        metadata_json=metadata_json,
        created_by_user_id=(user_id or "").strip() or None,
    )
    if existing_norms is not None:
        existing_norms.add(norm)
    return doc, None


def _parse_batch_create_items(raw: Any) -> list[dict[str, Any]]:
    if not isinstance(raw, list):
        return []
    items: list[dict[str, Any]] = []
    for item in raw:
        if not isinstance(item, dict):
            continue
        items.append(item)
        if len(items) >= _BATCH_MAX_SIZE:
            break
    return items


@document_control_bp.post("/api/document-control/documents/batch-create")
def api_document_control_batch_create():
    blocked = _require_feature()
    if blocked is not None:
        return blocked
    wall = login_wall()
    if wall is not None:
        return wall
    org_id, _ = _org_context()
    data = request.get_json(silent=True) or {}
    items = _parse_batch_create_items(data.get("items"))
    if not items:
        return jsonify({"message": "请提供要新增的记录"}), 400
    user_id = (session.get("user_id") or "").strip() or None
    confirm_sync = bool(data.get("confirmProjectCodeSync"))
    existing = _load_existing_docs_by_norm(org_id)
    batch_norms: set[str] = set()
    created = 0
    failed: list[dict[str, Any]] = []
    created_items: list[dict[str, Any]] = []
    if not confirm_sync:
        for item in items:
            pc = (item.get("projectCode") or "").strip()
            if not pc:
                continue
            blocked = _project_code_gate_response(
                org_id,
                pc,
                project_id=(item.get("projectId") or "").strip() or None,
                project_name=(item.get("projectName") or "").strip() or None,
                registered_country=(item.get("registeredCountry") or "").strip() or None,
                confirm_sync=False,
            )
            if blocked:
                return blocked
    for index, item in enumerate(items):
        doc_num = (item.get("documentNumber") or "").strip()
        norm = normalize_document_number(doc_num)
        row_ref = {
            "index": index,
            "documentNumber": doc_num or None,
            "title": (item.get("title") or "").strip() or None,
        }
        if norm and norm in existing:
            failed.append({**row_ref, "message": "受控编号已存在"})
            continue
        doc, err = _controlled_document_from_manual_item(
            org_id,
            item,
            user_id=user_id,
            existing_norms=batch_norms,
            confirm_sync=confirm_sync,
        )
        if err or not doc:
            failed.append({**row_ref, "message": err or "无法创建"})
            continue
        try:
            with db.session.begin_nested():
                db.session.add(doc)
                db.session.flush()
            link_controlled_document_to_page1_project(doc, organization_id=org_id)
        except IntegrityError:
            failed.append({**row_ref, "message": "受控编号已存在（并发冲突）"})
            continue
        existing[norm] = doc
        created += 1
        created_items.append(_serialize_doc(doc))
    if created:
        try:
            db.session.commit()
        except IntegrityError:
            db.session.rollback()
            return jsonify({"message": "批量新增失败：存在受控编号冲突"}), 409
    else:
        db.session.rollback()
    message = f"已新增 {created} 条"
    if failed:
        message += f"，失败 {len(failed)} 条"
    return jsonify(
        {
            "message": message,
            "created": created,
            "failed": failed,
            "items": created_items,
        }
    )


@document_control_bp.route(
    "/api/document-control/documents/<doc_id>", methods=["PATCH", "DELETE"]
)
def api_document_control_document_detail(doc_id: str):
    blocked = _require_feature()
    if blocked is not None:
        return blocked
    wall = login_wall()
    if wall is not None:
        return wall
    org_id, _ = _org_context()
    doc = _get_org_document(doc_id, org_id)
    if not doc:
        return jsonify({"message": "记录不存在"}), 404
    if request.method == "DELETE":
        db.session.delete(doc)
        db.session.commit()
        return jsonify({"message": "已删除"})
    data = request.get_json(silent=True) or {}
    if "documentNumber" in data:
        doc_num = (data.get("documentNumber") or "").strip()
        norm = normalize_document_number(doc_num)
        if not norm:
            return jsonify({"message": "请填写有效文件编号"}), 400
        if norm != doc.normalized_document_number:
            doc.document_number = doc_num
            doc.normalized_document_number = norm
    err = _apply_document_payload(doc, data)
    if err:
        return jsonify({"message": err}), 400
    if "projectCode" in data and (doc.project_code or "").strip():
        blocked = _project_code_gate_response(
            org_id,
            doc.project_code,
            project_id=doc.project_id,
            project_name=doc.project_name,
            registered_country=doc.registered_country,
            exclude_document_id=doc.id,
            confirm_sync=bool(data.get("confirmProjectCodeSync")),
        )
        if blocked:
            return blocked
    if doc.status == _DOC_STATUS_CONTROLLED:
        conflict_msg = controlled_number_conflict_message(
            org_id, doc.normalized_document_number, exclude_id=doc.id
        )
        if conflict_msg:
            return jsonify({"message": conflict_msg}), 409
    try:
        link_controlled_document_to_page1_project(doc, organization_id=org_id)
        db.session.commit()
    except IntegrityError:
        db.session.rollback()
        return jsonify({"message": "保存失败：受控编号冲突"}), 409
    return jsonify({"message": "已保存", "item": _serialize_doc(doc)})


_BATCH_EDIT_FIELDS = frozenset(
    {
        "status",
        "registrationSubmitted",
        "projectName",
        "projectCode",
        "registeredCountry",
        "sheetCategory",
    }
)
_BATCH_MAX_SIZE = 500


def _parse_batch_ids(raw: Any) -> list[str]:
    if not isinstance(raw, list):
        return []
    ids: list[str] = []
    for item in raw:
        value = str(item or "").strip()
        if value and value not in ids:
            ids.append(value)
        if len(ids) >= _BATCH_MAX_SIZE:
            break
    return ids


def _build_batch_update_payload(data: dict[str, Any]) -> tuple[dict[str, Any], Optional[str]]:
    payload: dict[str, Any] = {}
    if "status" in data:
        status = _parse_status_filter(str(data.get("status") or ""))
        if not status:
            return {}, "状态取值无效"
        payload["status"] = status
    if "registrationSubmitted" in data:
        payload["registrationSubmitted"] = bool(data.get("registrationSubmitted"))
    if "projectName" in data:
        payload["projectName"] = _format_scope_display(str(data.get("projectName") or ""))
    if "projectCode" in data:
        payload["projectCode"] = (data.get("projectCode") or "").strip()
    if "registeredCountry" in data:
        payload["registeredCountry"] = _format_scope_display(str(data.get("registeredCountry") or ""))
    if "sheetCategory" in data:
        payload["sheetCategory"] = (data.get("sheetCategory") or "").strip()
    unknown = set(data.keys()) - {"ids"} - _BATCH_EDIT_FIELDS
    if unknown:
        return {}, f"不支持批量修改字段：{', '.join(sorted(unknown))}"
    if not payload:
        return {}, "请至少指定一项要修改的内容"
    return payload, None


@document_control_bp.post("/api/document-control/documents/batch-update")
def api_document_control_batch_update():
    blocked = _require_feature()
    if blocked is not None:
        return blocked
    wall = login_wall()
    if wall is not None:
        return wall
    org_id, _ = _org_context()
    data = request.get_json(silent=True) or {}
    ids = _parse_batch_ids(data.get("ids"))
    if not ids:
        return jsonify({"message": "请先选择要修改的记录"}), 400
    payload, err = _build_batch_update_payload(data)
    if err:
        return jsonify({"message": err}), 400
    confirm_sync = bool(data.get("confirmProjectCodeSync"))
    docs = (
        ControlledDocument.query.filter(
            ControlledDocument.organization_id == org_id,
            ControlledDocument.id.in_(ids),
        )
        .all()
    )
    found_ids = {doc.id for doc in docs}
    missing = [doc_id for doc_id in ids if doc_id not in found_ids]
    updated = 0
    failed: list[dict[str, str]] = []
    for doc_id in missing:
        failed.append({"id": doc_id, "message": "记录不存在"})
    if "projectCode" in payload and (payload.get("projectCode") or "").strip() and not confirm_sync:
        for doc in docs:
            eff_name = (
                payload.get("projectName")
                if "projectName" in payload
                else doc.project_name
            )
            eff_country = (
                payload.get("registeredCountry")
                if "registeredCountry" in payload
                else doc.registered_country
            )
            blocked = _project_code_gate_response(
                org_id,
                payload["projectCode"],
                project_id=doc.project_id,
                project_name=eff_name,
                registered_country=eff_country,
                exclude_document_id=doc.id,
                confirm_sync=False,
            )
            if blocked:
                return blocked
    for doc in docs:
        next_status = (
            payload["status"]
            if "status" in payload
            else (doc.status or _DOC_STATUS_CONTROLLED)
        )
        if next_status == _DOC_STATUS_CONTROLLED:
            conflict_msg = controlled_number_conflict_message(
                org_id, doc.normalized_document_number, exclude_id=doc.id
            )
            if conflict_msg:
                failed.append({"id": doc.id, "message": conflict_msg})
                continue
        if "projectCode" in payload and (payload.get("projectCode") or "").strip():
            eff_name = (
                payload.get("projectName")
                if "projectName" in payload
                else doc.project_name
            )
            eff_country = (
                payload.get("registeredCountry")
                if "registeredCountry" in payload
                else doc.registered_country
            )
            gate = gate_project_code_save(
                org_id,
                payload["projectCode"],
                project_id=doc.project_id,
                project_name=eff_name,
                registered_country=eff_country,
                exclude_document_id=doc.id,
                confirm_sync=confirm_sync,
            )
            if gate:
                if gate[0] == "confirm":
                    return jsonify(gate[1]), 409
                failed.append({"id": doc.id, "message": gate[1]})
                continue
        apply_err = _apply_document_payload(doc, payload)
        if apply_err:
            failed.append({"id": doc.id, "message": apply_err})
            continue
        updated += 1
    if updated:
        try:
            db.session.commit()
        except IntegrityError:
            db.session.rollback()
            return jsonify({"message": "批量保存失败：存在受控编号冲突"}), 409
    else:
        db.session.rollback()
    message = f"已更新 {updated} 条"
    if failed:
        message += f"，失败 {len(failed)} 条"
    return jsonify({"message": message, "updated": updated, "failed": failed})


@document_control_bp.post("/api/document-control/documents/batch-delete")
def api_document_control_batch_delete():
    blocked = _require_feature()
    if blocked is not None:
        return blocked
    wall = login_wall()
    if wall is not None:
        return wall
    org_id, _ = _org_context()
    data = request.get_json(silent=True) or {}
    ids = _parse_batch_ids(data.get("ids"))
    if not ids:
        return jsonify({"message": "请先选择要删除的记录"}), 400
    docs = (
        ControlledDocument.query.filter(
            ControlledDocument.organization_id == org_id,
            ControlledDocument.id.in_(ids),
        )
        .all()
    )
    deleted = len(docs)
    for doc in docs:
        db.session.delete(doc)
    db.session.commit()
    missing = len(ids) - deleted
    message = f"已删除 {deleted} 条"
    if missing:
        message += f"，未找到 {missing} 条"
    return jsonify({"message": message, "deleted": deleted, "missing": missing})


@document_control_bp.get("/api/document-control/categories")
def api_document_control_categories():
    blocked = _require_feature()
    if blocked is not None:
        return blocked
    wall = login_wall()
    if wall is not None:
        return wall
    org_id, _ = _org_context()
    all_items = _list_sheet_categories(org_id)
    only_sheet = (request.args.get("sheetCategory") or "").strip()
    if only_sheet:
        q = _apply_document_filters(_documents_base_query(org_id))
        q = _apply_sheet_category_filter(q, only_sheet)
        counts = {only_sheet: int(q.count() or 0)}
        visible_items = [only_sheet] if counts.get(only_sheet, 0) > 0 else []
    else:
        counts = _filtered_category_counts(org_id)
        visible_items = [name for name in all_items if counts.get(name, 0) > 0]
    visible_items = _sort_sheet_categories(visible_items, org_id)
    sheet_order = _load_sheet_order_map().get(org_id) or []
    return jsonify(
        {
            "items": visible_items,
            "counts": counts,
            "allItems": all_items,
            "sheetOrder": sheet_order,
        }
    )


def _serialize_import_log(row: DocumentControlImportLog) -> dict[str, Any]:
    return {
        "id": row.id,
        "batchId": row.import_batch_id,
        "eventType": row.event_type,
        "documentNumber": row.document_number,
        "sheetName": row.sheet_name,
        "rowIndex": row.row_index,
        "reason": row.reason,
        "controlledDocumentId": row.controlled_document_id,
        "createdAt": row.created_at.isoformat() if row.created_at else None,
    }


def _import_logs_base_query(org_id: str):
    return DocumentControlImportLog.query.filter_by(organization_id=org_id)


_IMPORT_LOG_MAX_BATCHES = 100


def _import_log_batch_summaries(
    org_id: str, *, batch_id: Optional[str] = None
) -> list[dict[str, Any]]:
    q = _import_logs_base_query(org_id)
    if batch_id:
        q = q.filter(DocumentControlImportLog.import_batch_id == batch_id)
    q = q.filter(DocumentControlImportLog.import_batch_id.isnot(None))
    q = q.filter(DocumentControlImportLog.import_batch_id != "")
    q = (
        q.with_entities(
            DocumentControlImportLog.import_batch_id.label("batch_id"),
            func.min(DocumentControlImportLog.created_at).label("started_at"),
            func.count().label("total"),
            func.sum(
                case((DocumentControlImportLog.event_type == "import_success", 1), else_=0)
            ).label("success"),
            func.sum(
                case((DocumentControlImportLog.event_type == "import_update", 1), else_=0)
            ).label("updated"),
            func.sum(
                case((DocumentControlImportLog.event_type == "import_skip", 1), else_=0)
            ).label("skip"),
            func.sum(
                case((DocumentControlImportLog.event_type == "import_fail", 1), else_=0)
            ).label("fail"),
            func.sum(
                case((DocumentControlImportLog.event_type == "registration_link", 1), else_=0)
            ).label("registrationLink"),
        )
        .group_by(DocumentControlImportLog.import_batch_id)
        .order_by(func.min(DocumentControlImportLog.created_at).desc())
    )
    if not batch_id:
        q = q.limit(_IMPORT_LOG_MAX_BATCHES)
    rows = q.all()
    summaries: list[dict[str, Any]] = []
    for row in rows:
        summaries.append(
            {
                "batchId": row.batch_id,
                "startedAt": row.started_at.isoformat() if row.started_at else None,
                "total": int(row.total or 0),
                "success": int(row.success or 0),
                "updated": int(row.updated or 0),
                "skip": int(row.skip or 0),
                "fail": int(row.fail or 0),
                "registrationLink": int(row.registrationLink or 0),
            }
        )
    return summaries


@document_control_bp.get("/api/document-control/import/logs")
def api_document_control_import_logs():
    blocked = _require_feature()
    if blocked is not None:
        return blocked
    wall = login_wall()
    if wall is not None:
        return wall
    org_id, _ = _org_context()
    batch_id = (request.args.get("batchId") or "").strip()
    event_type = (request.args.get("eventType") or "").strip()
    page = max(1, int(request.args.get("page") or 1))
    page_size = min(500, max(10, int(request.args.get("pageSize") or 100)))
    q = _import_logs_base_query(org_id)
    if batch_id:
        q = q.filter(DocumentControlImportLog.import_batch_id == batch_id)
    if event_type:
        q = q.filter(DocumentControlImportLog.event_type == event_type)
    total = q.count()
    rows = (
        q.order_by(
            DocumentControlImportLog.created_at.desc(),
            DocumentControlImportLog.id.desc(),
        )
        .offset((page - 1) * page_size)
        .limit(page_size)
        .all()
    )
    batches = _import_log_batch_summaries(org_id, batch_id=batch_id or None)
    return jsonify(
        {
            "items": [_serialize_import_log(row) for row in rows],
            "batches": batches,
            "total": total,
            "page": page,
            "pageSize": page_size,
        }
    )


@document_control_bp.get("/api/document-control/search")
def api_document_control_search():
    return api_document_control_documents()


@document_control_bp.route("/api/document-control/schemes", methods=["GET", "POST"])
def api_document_control_schemes():
    blocked = _require_feature()
    if blocked is not None:
        return blocked
    wall = login_wall()
    if wall is not None:
        return wall
    org_id, _ = _org_context()
    if request.method == "GET":
        rows = (
            NumberingScheme.query.filter_by(organization_id=org_id)
            .order_by(NumberingScheme.created_at.desc())
            .all()
        )
        return jsonify({"items": [_serialize_scheme(x) for x in rows]})
    data = request.get_json(force=True) or {}
    name = (data.get("name") or "").strip()
    doc_type = (data.get("docTypeCode") or "").strip().upper()
    if not name or not doc_type:
        return jsonify({"message": "name 与 docTypeCode 必填"}), 400
    row = NumberingScheme(
        organization_id=org_id,
        name=name,
        doc_type_code=doc_type,
        pattern_regex=(data.get("patternRegex") or "").strip() or None,
        render_template=(data.get("renderTemplate") or "").strip()
        or "{prefix}-{type}-{seq:03d}",
        prefix_source=(data.get("prefixSource") or "fixed").strip() or "fixed",
        fixed_prefix=(data.get("fixedPrefix") or "").strip() or None,
        seq_scope=(data.get("seqScope") or "per_company").strip() or "per_company",
        seq_start=max(1, int(data.get("seqStart") or 1)),
        seq_pad=max(1, int(data.get("seqPad") or 3)),
        is_active=bool(data.get("isActive", True)),
        kb_rule_excerpt=(data.get("kbRuleExcerpt") or "").strip() or None,
    )
    db.session.add(row)
    db.session.commit()
    return jsonify({"message": "已创建", "item": _serialize_scheme(row)})


@document_control_bp.route("/api/document-control/schemes/<scheme_id>", methods=["PATCH", "DELETE"])
def api_document_control_scheme_update(scheme_id: str):
    blocked = _require_feature()
    if blocked is not None:
        return blocked
    wall = login_wall()
    if wall is not None:
        return wall
    org_id, _ = _org_context()
    row = NumberingScheme.query.filter_by(id=scheme_id, organization_id=org_id).first()
    if not row:
        return jsonify({"message": "规则不存在"}), 404
    if request.method == "DELETE":
        db.session.delete(row)
        db.session.commit()
        return jsonify({"message": "已删除"})
    data = request.get_json(force=True) or {}
    for attr, key in (
        ("name", "name"),
        ("doc_type_code", "docTypeCode"),
        ("pattern_regex", "patternRegex"),
        ("render_template", "renderTemplate"),
        ("prefix_source", "prefixSource"),
        ("fixed_prefix", "fixedPrefix"),
        ("seq_scope", "seqScope"),
        ("kb_rule_excerpt", "kbRuleExcerpt"),
    ):
        if key in data:
            setattr(row, attr, (data.get(key) or "").strip() or None)
    if "seqStart" in data:
        row.seq_start = max(1, int(data.get("seqStart") or 1))
    if "seqPad" in data:
        row.seq_pad = max(1, int(data.get("seqPad") or 3))
    if "isActive" in data:
        row.is_active = bool(data.get("isActive"))
    db.session.add(row)
    db.session.commit()
    return jsonify({"message": "已更新", "item": _serialize_scheme(row)})


def _apply_title_en_to_doc(doc: ControlledDocument, title_en: Optional[str]) -> None:
    from sqlalchemy.orm.attributes import flag_modified

    te = (title_en or "").strip()
    if not te:
        return
    doc.title_en = te
    meta = dict(doc.metadata_json) if isinstance(doc.metadata_json, dict) else {}
    meta["titleEn"] = te
    doc.metadata_json = meta
    flag_modified(doc, "metadata_json")


def _prepare_issue_number_inputs(
    org_id: str,
    collection: str,
    scheme: NumberingScheme,
    cfg: Optional[dict[str, Any]],
    data: dict[str, Any],
    *,
    session_cache: Optional[dict[str, tuple[str, str]]] = None,
) -> dict[str, Any]:
    title = (data.get("title") or "").strip()
    project_code = (data.get("projectCode") or "").strip() or None
    project_id = (data.get("projectId") or "").strip() or None
    project_name = (data.get("projectName") or "").strip() or None
    subtype_from_title = bool((cfg or {}).get("subtypeFromTitle")) or _scheme_uses_subtype_template(scheme)
    force_new = bool(data.get("forceNew"))
    confirm_same = bool(data.get("confirmSameDocument"))
    manual_subtype = (data.get("subtype") or data.get("docSubtype") or "").strip() or None
    prefix = scheme_allocation_prefix(scheme, project_code)
    existing = None
    if title and not force_new:
        existing = find_existing_controlled_doc_by_title(
            organization_id=org_id,
            prefix=prefix,
            title=title,
            project_id=project_id,
        )
    title_en = None
    title_en_source = None
    if subtype_from_title and title:
        try:
            title_en, title_en_source = resolve_title_en_for_issue(
                title,
                org_id=org_id,
                collection=collection,
                cached_title_en=(data.get("titleEn") or "").strip() or None,
                session_cache=session_cache,
            )
        except ValueError:
            if not existing:
                raise
    return {
        "title": title,
        "project_code": project_code,
        "project_id": project_id,
        "project_name": project_name,
        "subtype_from_title": subtype_from_title,
        "force_new": force_new,
        "confirm_same": confirm_same,
        "manual_subtype": manual_subtype,
        "title_en": title_en,
        "title_en_source": title_en_source,
        "prefix": prefix,
        "existing": existing,
    }


def _duplicate_issue_message(title: str, existing: ControlledDocument) -> str:
    return (
        f"台账中已有名称包含「{title}」的受控文件「{existing.title or '-'}」，"
        f"编号 {existing.document_number or '-'}。请确认是否为同一份文件。"
    )


def _allocate_preview_item(
    org_id: str,
    collection: str,
    scheme: NumberingScheme,
    cfg: Optional[dict[str, Any]],
    data: dict[str, Any],
    *,
    session_cache: Optional[dict[str, tuple[str, str]]] = None,
) -> dict[str, Any]:
    title = (data.get("title") or "").strip()
    if not title:
        return {"title": "", "error": "文件名称不能为空"}
    try:
        ctx = _prepare_issue_number_inputs(
            org_id, collection, scheme, cfg, data, session_cache=session_cache
        )
    except ValueError as exc:
        return {"title": title, "error": str(exc)}
    existing = ctx["existing"]
    if existing:
        return {
            "title": title,
            "duplicateTitle": True,
            "existingDocument": _serialize_doc(existing),
            "titleEn": ctx["title_en"],
            "titleEnSource": ctx["title_en_source"],
            "message": _duplicate_issue_message(title, existing),
        }
    try:
        info = preview_next_number(
            organization_id=org_id,
            scheme=scheme,
            project_code=ctx["project_code"],
            project_id=ctx["project_id"],
            project_name=ctx["project_name"],
            subtype=ctx["manual_subtype"],
            title=ctx["title"],
            title_en=ctx["title_en"],
            subtype_from_title=ctx["subtype_from_title"],
        )
    except SubtypeChoiceRequired as exc:
        return {
            "title": title,
            "titleEn": ctx["title_en"],
            "titleEnSource": ctx["title_en_source"],
            "subtypeChoiceRequired": True,
            "subtypeChoices": exc.choices,
            "message": str(exc),
        }
    except ValueError as exc:
        return {"title": title, "titleEn": ctx["title_en"], "titleEnSource": ctx["title_en_source"], "error": str(exc)}
    if ctx["title_en"] and not info.get("title_en"):
        info["title_en"] = ctx["title_en"]
    return {
        "title": title,
        "titleEn": ctx["title_en"],
        "titleEnSource": ctx["title_en_source"],
        "duplicateTitle": False,
        "preview": info,
        "documentNumber": info.get("document_number"),
    }


def _allocate_apply_item(
    org_id: str,
    collection: str,
    scheme: NumberingScheme,
    cfg: Optional[dict[str, Any]],
    data: dict[str, Any],
    *,
    sheet_cat: str,
    project_name: Optional[str],
    session_cache: Optional[dict[str, tuple[str, str]]] = None,
) -> dict[str, Any]:
    title = (data.get("title") or "").strip()
    if not title:
        return {"title": "", "ok": False, "error": "文件名称不能为空"}
    try:
        ctx = _prepare_issue_number_inputs(
            org_id, collection, scheme, cfg, data, session_cache=session_cache
        )
    except ValueError as exc:
        return {"title": title, "ok": False, "error": str(exc)}

    existing = ctx["existing"]
    if existing and ctx["confirm_same"]:
        if ctx["title_en"]:
            _apply_title_en_to_doc(existing, ctx["title_en"])
            db.session.add(existing)
        return {
            "title": title,
            "ok": True,
            "duplicateTitle": True,
            "confirmedSameDocument": True,
            "document": _serialize_doc(existing),
            "message": f"该文件已在台账中，编号 {existing.document_number or '-'}",
        }
    if existing and not ctx["force_new"]:
        return {
            "title": title,
            "ok": False,
            "duplicateTitle": True,
            "existingDocument": _serialize_doc(existing),
            "titleEn": ctx["title_en"],
            "titleEnSource": ctx["title_en_source"],
            "message": _duplicate_issue_message(title, existing),
        }
    try:
        allocation = reserve_number(
            organization_id=org_id,
            scheme=scheme,
            requested_title=title,
            project_id=ctx["project_id"],
            project_code=ctx["project_code"],
            project_name=ctx["project_name"],
            user_id=(session.get("user_id") or "").strip() or None,
            reserved_minutes=int(data.get("reservedMinutes") or 30),
            subtype=ctx["manual_subtype"],
            subtype_from_title=ctx["subtype_from_title"],
            title_en=ctx["title_en"],
        )
        doc = issue_number(
            organization_id=org_id,
            allocation=allocation,
            version=(data.get("version") or "").strip() or None,
            title=title,
            source="allocated",
            project_id=ctx["project_id"],
            project_code=ctx["project_code"],
            user_id=(session.get("user_id") or "").strip() or None,
            title_en=ctx["title_en"],
        )
        if sheet_cat:
            doc.sheet_category = sheet_cat
        if project_name:
            doc.project_name = project_name
        reg_country = _normalize_scope_field_value(data.get("registeredCountry"))
        if reg_country:
            doc.registered_country = reg_country
        if ctx["title_en"]:
            _apply_title_en_to_doc(doc, ctx["title_en"])
            if ctx["title_en_source"] == "translated":
                persist_title_en_cache(org_id, title, ctx["title_en"], source="translated")
        db.session.add(doc)
        return {
            "title": title,
            "ok": True,
            "document": _serialize_doc(doc),
            "documentNumber": doc.document_number,
            "titleEn": ctx["title_en"],
            "titleEnSource": ctx["title_en_source"],
            "message": f"已申请编号：{doc.document_number}",
        }
    except SubtypeChoiceRequired as exc:
        db.session.rollback()
        return {
            "title": title,
            "ok": False,
            "subtypeChoiceRequired": True,
            "subtypeChoices": exc.choices,
            "titleEn": ctx["title_en"],
            "titleEnSource": ctx["title_en_source"],
            "message": str(exc),
        }
    except ValueError as exc:
        db.session.rollback()
        return {"title": title, "ok": False, "error": str(exc)}


def _parse_batch_titles(data: dict[str, Any]) -> list[str]:
    items = data.get("items")
    if isinstance(items, list) and items:
        out: list[str] = []
        seen: set[str] = set()
        for row in items:
            if not isinstance(row, dict):
                continue
            t = (row.get("title") or "").strip()
            if not t:
                continue
            key = t.casefold()
            if key in seen:
                continue
            seen.add(key)
            out.append(t)
        return out
    raw = (data.get("titlesText") or data.get("titles") or "")
    if isinstance(raw, list):
        lines = [str(x) for x in raw]
    else:
        lines = str(raw or "").splitlines()
    out = []
    seen: set[str] = set()
    for line in lines:
        t = (line or "").strip()
        if not t or t.startswith("#"):
            continue
        key = t.casefold()
        if key in seen:
            continue
        seen.add(key)
        out.append(t)
    return out


@document_control_bp.post("/api/document-control/allocate/preview")
def api_document_control_allocate_preview():
    blocked = _require_feature()
    if blocked is not None:
        return blocked
    wall = login_wall()
    if wall is not None:
        return wall
    org_id, collection = _org_context()
    data = request.get_json(force=True) or {}
    release_expired_number_reservations(org_id)
    scheme_id = (data.get("schemeId") or "").strip()
    sheet_category = (data.get("sheetCategory") or "").strip()
    scheme, cfg, err = resolve_scheme_for_issue(
        org_id,
        sheet_category=sheet_category,
        scheme_id=scheme_id or None,
    )
    if not scheme:
        return jsonify({"message": err or "未找到编号规则"}), 400
    try:
        ctx = _prepare_issue_number_inputs(org_id, collection, scheme, cfg, data)
    except ValueError as exc:
        return jsonify({"message": str(exc)}), 400
    if not ctx["title"]:
        return jsonify({"message": "请填写文件名称"}), 400
    if ctx["subtype_from_title"] and (scheme.prefix_source or "") == "from_project_code" and not ctx["project_code"]:
        return jsonify({"message": "请选择项目或填写项目编号"}), 400
    existing = ctx["existing"]
    if existing:
        try:
            db.session.commit()
        except Exception:
            db.session.rollback()
        return jsonify(
            {
                "duplicateTitle": True,
                "existingDocument": _serialize_doc(existing),
                "titleEn": ctx["title_en"],
                "titleEnSource": ctx["title_en_source"],
                "message": (
                    f"台账中已有名称包含「{ctx['title']}」的受控文件「{existing.title or '-'}」，"
                    f"编号 {existing.document_number or '-'}。请确认是否为同一份文件。"
                ),
            }
        )
    try:
        info = preview_next_number(
            organization_id=org_id,
            scheme=scheme,
            project_code=ctx["project_code"],
            project_id=ctx["project_id"],
            project_name=ctx["project_name"],
            subtype=ctx["manual_subtype"],
            title=ctx["title"],
            title_en=ctx["title_en"],
            subtype_from_title=ctx["subtype_from_title"],
        )
    except SubtypeChoiceRequired as exc:
        try:
            db.session.commit()
        except Exception:
            db.session.rollback()
        return jsonify(
            {
                "subtypeChoiceRequired": True,
                "subtypeChoices": exc.choices,
                "titleEn": ctx["title_en"],
                "titleEnSource": ctx["title_en_source"],
                "message": str(exc),
            }
        )
    except ValueError as exc:
        return jsonify({"message": str(exc)}), 400
    except Exception as exc:
        current_app.logger.exception("document-control allocate preview failed")
        return jsonify({"message": "预览编号失败，请稍后重试"}), 500
    if ctx["title_en"] and not info.get("title_en"):
        info["title_en"] = ctx["title_en"]
    try:
        db.session.commit()
    except Exception:
        db.session.rollback()
    return jsonify(
        {
            "item": info,
            "scheme": _serialize_scheme(scheme),
            "titleEn": ctx["title_en"],
            "titleEnSource": ctx["title_en_source"],
        }
    )


def _page1_projects_for_issue(org_id: str) -> list[dict[str, Any]]:
    """页面1 进行中项目，附带任务录入里常用的项目编号。"""
    from webapp.authz import filter_projects_by_organization, project_in_scope, rbac_enforced
    from webapp.models import Project, UploadRecord

    rows = (
        Project.query.filter(Project.status == Project.STATUS_ACTIVE)
        .order_by(Project.name.asc())
        .all()
    )
    if rbac_enforced():
        rows = [p for p in rows if project_in_scope(p)]
    rows = filter_projects_by_organization(rows)
    pid_list = [p.id for p in rows if p.id]
    code_by_pid: dict[str, str] = {}
    for p in rows:
        pc = (getattr(p, "project_code", None) or "").strip()
        if p.id and pc:
            code_by_pid[p.id] = pc
    if pid_list:
        missing = [x for x in pid_list if x not in code_by_pid]
        if missing:
            for ur in (
                UploadRecord.query.filter(UploadRecord.project_id.in_(missing))
                .order_by(UploadRecord.updated_at.desc())
                .all()
            ):
                pid = (ur.project_id or "").strip()
                pc = (ur.project_code or "").strip()
                if pid and pc and pid not in code_by_pid:
                    code_by_pid[pid] = pc
    items: list[dict[str, Any]] = []
    for p in rows:
        name = (p.name or "").strip()
        code = code_by_pid.get(p.id or "", "") or ""
        label = name if not code else f"{name}（{code}）"
        items.append(
            {
                "id": p.id,
                "name": name,
                "projectCode": code,
                "projectName": name,
                "registeredCountry": (getattr(p, "registered_country", None) or "").strip() or None,
                "label": label,
            }
        )
    return items


@document_control_bp.get("/api/document-control/allocate/options")
def api_document_control_allocate_options():
    blocked = _require_feature()
    if blocked is not None:
        return blocked
    wall = login_wall()
    if wall is not None:
        return wall
    org_id, _ = _org_context()
    return jsonify(
        {
            "categories": enrich_issue_categories(org_id),
            "projects": _page1_projects_for_issue(org_id),
        }
    )


@document_control_bp.post("/api/document-control/allocate/reserve")
def api_document_control_allocate_reserve():
    blocked = _require_feature()
    if blocked is not None:
        return blocked
    wall = login_wall()
    if wall is not None:
        return wall
    org_id, _ = _org_context()
    data = request.get_json(force=True) or {}
    scheme_id = (data.get("schemeId") or "").strip()
    scheme = NumberingScheme.query.filter_by(id=scheme_id, organization_id=org_id).first()
    if not scheme:
        return jsonify({"message": "规则不存在"}), 404
    try:
        row = reserve_number(
            organization_id=org_id,
            scheme=scheme,
            requested_title=(data.get("title") or "").strip(),
            project_id=(data.get("projectId") or "").strip() or None,
            project_code=(data.get("projectCode") or "").strip() or None,
            user_id=(session.get("user_id") or "").strip() or None,
            reserved_minutes=int(data.get("reservedMinutes") or 30),
        )
        db.session.commit()
        return jsonify(
            {
                "message": "已预留",
                "item": {
                    "id": row.id,
                    "allocatedNumber": row.allocated_number,
                    "status": row.status,
                    "reservedUntil": row.reserved_until.isoformat()
                    if row.reserved_until
                    else None,
                },
            }
        )
    except ValueError as exc:
        db.session.rollback()
        return jsonify({"message": str(exc)}), 409


@document_control_bp.post("/api/document-control/allocate/issue")
def api_document_control_allocate_issue():
    blocked = _require_feature()
    if blocked is not None:
        return blocked
    wall = login_wall()
    if wall is not None:
        return wall
    org_id, _ = _org_context()
    data = request.get_json(force=True) or {}
    allocation_id = (data.get("allocationId") or "").strip()
    allocation = NumberAllocation.query.filter_by(
        id=allocation_id, organization_id=org_id
    ).first()
    if not allocation:
        return jsonify({"message": "预留记录不存在"}), 404
    doc = issue_number(
        organization_id=org_id,
        allocation=allocation,
        version=(data.get("version") or "").strip() or None,
        title=(data.get("title") or "").strip() or allocation.requested_title or "未命名文件",
        source=(data.get("source") or "").strip() or "allocated",
        upload_record_id=(data.get("uploadRecordId") or "").strip() or None,
        project_id=(data.get("projectId") or "").strip() or None,
        project_code=(data.get("projectCode") or "").strip() or None,
        user_id=(session.get("user_id") or "").strip() or None,
    )
    db.session.commit()
    return jsonify({"message": "已发放", "document": _serialize_doc(doc)})


@document_control_bp.post("/api/document-control/allocate/apply")
def api_document_control_allocate_apply():
    """预留编号并写入受控台账（一步完成）。"""
    blocked = _require_feature()
    if blocked is not None:
        return blocked
    wall = login_wall()
    if wall is not None:
        return wall
    org_id, collection = _org_context()
    data = request.get_json(force=True) or {}
    release_expired_number_reservations(org_id)
    title = (data.get("title") or "").strip()
    if not title:
        return jsonify({"message": "请填写文件名称"}), 400
    sheet_category = (data.get("sheetCategory") or "").strip()
    scheme_id = (data.get("schemeId") or "").strip()
    scheme, cfg, err = resolve_scheme_for_issue(
        org_id,
        sheet_category=sheet_category,
        scheme_id=scheme_id or None,
    )
    if not scheme:
        return jsonify({"message": err or "该分类不支持自动取号"}), 400
    tmpl = (scheme.render_template or "").strip()
    if not scheme.is_active or not tmpl or "{seq" not in tmpl:
        hint = (scheme.kb_rule_excerpt or "").strip() or "该文件类型须按《文件控制程序》手工编号"
        return jsonify({"message": f"该类型不支持自动取号：{hint}"}), 400
    project_name = (data.get("projectName") or "").strip() or None
    sheet_cat = sheet_category or ((cfg or {}).get("sheetCategory") or "").strip()
    try:
        ctx = _prepare_issue_number_inputs(org_id, collection, scheme, cfg, data)
    except ValueError as exc:
        return jsonify({"message": str(exc)}), 400
    if (scheme.prefix_source or "") == "from_project_code" and not ctx["project_code"]:
        return jsonify({"message": "请选择项目或填写项目编号"}), 400

    existing = ctx["existing"]
    if existing and ctx["confirm_same"]:
        if ctx["title_en"]:
            _apply_title_en_to_doc(existing, ctx["title_en"])
            db.session.add(existing)
            db.session.commit()
        return jsonify(
            {
                "message": f"该文件已在台账中，编号 {existing.document_number or '-'}",
                "duplicateTitle": True,
                "confirmedSameDocument": True,
                "document": _serialize_doc(existing),
            }
        )
    if existing and not ctx["force_new"]:
        return jsonify(
            {
                "duplicateTitle": True,
                "existingDocument": _serialize_doc(existing),
                "titleEn": ctx["title_en"],
                "titleEnSource": ctx["title_en_source"],
                "message": (
                    f"台账中已有名称包含「{ctx['title']}」的受控文件「{existing.title or '-'}」，"
                    f"编号 {existing.document_number or '-'}。请确认是否为同一份文件。"
                ),
            }
        ), 409

    try:
        allocation = reserve_number(
            organization_id=org_id,
            scheme=scheme,
            requested_title=title,
            project_id=ctx["project_id"],
            project_code=ctx["project_code"],
            project_name=ctx["project_name"],
            user_id=(session.get("user_id") or "").strip() or None,
            reserved_minutes=int(data.get("reservedMinutes") or 30),
            subtype=ctx["manual_subtype"],
            subtype_from_title=ctx["subtype_from_title"],
            title_en=ctx["title_en"],
        )
        doc = issue_number(
            organization_id=org_id,
            allocation=allocation,
            version=(data.get("version") or "").strip() or None,
            title=title,
            source="allocated",
            project_id=ctx["project_id"],
            project_code=ctx["project_code"],
            user_id=(session.get("user_id") or "").strip() or None,
            title_en=ctx["title_en"],
        )
        if sheet_cat:
            doc.sheet_category = sheet_cat
        if project_name:
            doc.project_name = project_name
        reg_country = _normalize_scope_field_value(data.get("registeredCountry"))
        if reg_country:
            doc.registered_country = reg_country
        if ctx["title_en"]:
            _apply_title_en_to_doc(doc, ctx["title_en"])
            if ctx["title_en_source"] == "translated":
                persist_title_en_cache(org_id, title, ctx["title_en"], source="translated")
        db.session.add(doc)
        db.session.commit()
        return jsonify(
            {
                "message": f"已申请编号：{doc.document_number}",
                "document": _serialize_doc(doc),
                "allocationId": allocation.id,
                "titleEn": ctx["title_en"],
                "titleEnSource": ctx["title_en_source"],
            }
        )
    except SubtypeChoiceRequired as exc:
        db.session.rollback()
        return jsonify(
            {
                "subtypeChoiceRequired": True,
                "subtypeChoices": exc.choices,
                "titleEn": ctx["title_en"],
                "titleEnSource": ctx["title_en_source"],
                "message": str(exc),
            }
        ), 400
    except ValueError as exc:
        db.session.rollback()
        msg = str(exc)
        status = 409 if "已被占用" in msg else 400
        return jsonify({"message": msg}), status
    except Exception:
        db.session.rollback()
        current_app.logger.exception("document-control allocate apply failed")
        return jsonify({"message": "申请编号失败，请稍后重试"}), 500


@document_control_bp.post("/api/document-control/allocate/batch/preview")
def api_document_control_allocate_batch_preview():
    """批量预览：台账/缓存匹配英文名，仅未命中才调 AI；批量判重。"""
    blocked = _require_feature()
    if blocked is not None:
        return blocked
    wall = login_wall()
    if wall is not None:
        return wall
    org_id, collection = _org_context()
    data = request.get_json(force=True) or {}
    release_expired_number_reservations(org_id)
    titles = _parse_batch_titles(data)
    if not titles:
        return jsonify({"message": "请填写至少一个文件名称"}), 400
    if len(titles) > 100:
        return jsonify({"message": "单次批量最多 100 个文件名称"}), 400
    sheet_category = (data.get("sheetCategory") or "").strip()
    scheme, cfg, err = resolve_scheme_for_issue(
        org_id,
        sheet_category=sheet_category,
        scheme_id=(data.get("schemeId") or "").strip() or None,
    )
    if not scheme:
        return jsonify({"message": err or "未找到编号规则"}), 400
    if (scheme.prefix_source or "") == "from_project_code" and not (data.get("projectCode") or "").strip():
        return jsonify({"message": "请选择项目或填写项目编号"}), 400
    base = {
        "sheetCategory": sheet_category,
        "schemeId": scheme.id,
        "projectId": (data.get("projectId") or "").strip() or None,
        "projectCode": (data.get("projectCode") or "").strip() or None,
        "projectName": (data.get("projectName") or "").strip() or None,
        "registeredCountry": (data.get("registeredCountry") or "").strip() or None,
    }
    session_cache: dict[str, tuple[str, str]] = {}
    items: list[dict[str, Any]] = []
    dup_count = 0
    err_count = 0
    ready_count = 0
    subtype_choice_count = 0
    for title in titles:
        row = _allocate_preview_item(
            org_id,
            collection,
            scheme,
            cfg,
            {**base, "title": title},
            session_cache=session_cache,
        )
        items.append(row)
        if row.get("error"):
            err_count += 1
        elif row.get("duplicateTitle"):
            dup_count += 1
        elif row.get("subtypeChoiceRequired"):
            subtype_choice_count += 1
        else:
            ready_count += 1
    try:
        db.session.commit()
    except Exception:
        db.session.rollback()
        current_app.logger.exception("document-control batch preview cache commit failed")
    return jsonify(
        {
            "items": items,
            "total": len(items),
            "readyCount": ready_count,
            "duplicateCount": dup_count,
            "errorCount": err_count,
            "subtypeChoiceCount": subtype_choice_count,
        }
    )


@document_control_bp.post("/api/document-control/allocate/batch/apply")
def api_document_control_allocate_batch_apply():
    """批量申请编号：跳过判重项（除非 forceNew）；支持 confirmSameDocument。"""
    blocked = _require_feature()
    if blocked is not None:
        return blocked
    wall = login_wall()
    if wall is not None:
        return wall
    org_id, collection = _org_context()
    data = request.get_json(force=True) or {}
    release_expired_number_reservations(org_id)
    raw_items = data.get("items")
    if not isinstance(raw_items, list) or not raw_items:
        return jsonify({"message": "请提供 items 列表"}), 400
    if len(raw_items) > 100:
        return jsonify({"message": "单次批量最多 100 条"}), 400
    sheet_category = (data.get("sheetCategory") or "").strip()
    scheme, cfg, err = resolve_scheme_for_issue(
        org_id,
        sheet_category=sheet_category,
        scheme_id=(data.get("schemeId") or "").strip() or None,
    )
    if not scheme:
        return jsonify({"message": err or "该分类不支持自动取号"}), 400
    tmpl = (scheme.render_template or "").strip()
    if not scheme.is_active or not tmpl or "{seq" not in tmpl:
        hint = (scheme.kb_rule_excerpt or "").strip() or "该文件类型须按《文件控制程序》手工编号"
        return jsonify({"message": f"该类型不支持自动取号：{hint}"}), 400
    if (scheme.prefix_source or "") == "from_project_code" and not (data.get("projectCode") or "").strip():
        return jsonify({"message": "请选择项目或填写项目编号"}), 400
    sheet_cat = sheet_category or ((cfg or {}).get("sheetCategory") or "").strip()
    project_name = (data.get("projectName") or "").strip() or None
    base = {
        "sheetCategory": sheet_category,
        "schemeId": scheme.id,
        "projectId": (data.get("projectId") or "").strip() or None,
        "projectCode": (data.get("projectCode") or "").strip() or None,
        "projectName": project_name,
        "registeredCountry": (data.get("registeredCountry") or "").strip() or None,
    }
    session_cache: dict[str, tuple[str, str]] = {}
    results: list[dict[str, Any]] = []
    ok_count = 0
    skip_count = 0
    fail_count = 0
    for raw in raw_items:
        if not isinstance(raw, dict):
            continue
        title = (raw.get("title") or "").strip()
        if not title:
            continue
        if raw.get("skip"):
            skip_count += 1
            results.append({"title": title, "ok": False, "skipped": True, "message": "已跳过"})
            continue
        item_data = {
            **base,
            "title": title,
            "version": (raw.get("version") or "").strip() or None,
            "titleEn": (raw.get("titleEn") or "").strip() or None,
            "forceNew": bool(raw.get("forceNew")),
            "confirmSameDocument": bool(raw.get("confirmSameDocument")),
        }
        row = _allocate_apply_item(
            org_id,
            collection,
            scheme,
            cfg,
            item_data,
            sheet_cat=sheet_cat,
            project_name=project_name,
            session_cache=session_cache,
        )
        results.append(row)
        if row.get("ok"):
            ok_count += 1
            try:
                db.session.commit()
            except Exception:
                db.session.rollback()
                row["ok"] = False
                row["error"] = "写入失败，请重试"
                ok_count -= 1
                fail_count += 1
        elif row.get("duplicateTitle") and not row.get("confirmedSameDocument"):
            skip_count += 1
        else:
            fail_count += 1
    return jsonify(
        {
            "items": results,
            "okCount": ok_count,
            "skipCount": skip_count,
            "failCount": fail_count,
            "message": f"成功 {ok_count} 条，跳过 {skip_count} 条，失败 {fail_count} 条",
        }
    )


def _excel_import_header_map() -> dict[str, str]:
    return {
        "文件编号": "documentNumber",
        "编号": "documentNumber",
        "文件号": "documentNumber",
        "受控编号": "documentNumber",
        "文档编号": "documentNumber",
        "document number": "documentNumber",
        "documentnumber": "documentNumber",
        "文件名称": "title",
        "名称": "title",
        "文档名称": "title",
        "文件名": "title",
        "文件名称（中文）": "title",
        "title": "title",
        "英文名": "titleEn",
        "文件名称（英文）": "titleEn",
        "版本号": "version",
        "版本": "version",
        "version": "version",
        "文件类型": "docTypeCode",
        "类型": "docTypeCode",
        "项目编号": "projectCode",
        "项目号": "projectCode",
        "project code": "projectCode",
        "projectcode": "projectCode",
        "project no": "projectCode",
        "projectno": "projectCode",
        "所属项目": "projectName",
        "注册国家": "registeredCountry",
        "状态": "lifecycleStatus",
        "文件状态": "lifecycleStatus",
        "受控状态": "lifecycleStatus",
        "status": "lifecycleStatus",
    }


_EXCEL_VOIDED_STATUS_MARKERS = ("作废", "废止", "失效", "注销", "已作废", "已废止")
_EXCEL_CONTROLLED_STATUS_VALUES = frozenset(
    {
        "受控",
        "有效",
        "现行",
        "生效",
        "在用",
        "active",
        "controlled",
        "effective",
    }
)
_SHEET_ORDER_CONFIG_KEY = "DOCUMENT_CONTROL_SHEET_ORDER_BY_ORG"


def _normalize_sheet_tab_name(sheet_name: str) -> str:
    return (sheet_name or "").strip().replace(" ", "")


def _is_catalog_sheet(sheet_name: str) -> bool:
    """清单索引 Sheet：仅匹配「目录」「目录1」「目录2」等，不参与导入与操作日志。"""
    compact = _normalize_sheet_tab_name(sheet_name)
    if not compact:
        return False
    if compact == "目录":
        return True
    if compact.startswith("目录") and compact[2:].isdigit():
        return True
    return False


def _is_registration_sheet(sheet_name: str) -> bool:
    compact = _normalize_sheet_tab_name(sheet_name)
    return compact == _REGISTRATION_SHEET_NAME or compact.startswith(_REGISTRATION_SHEET_NAME)


def _is_dhf_sheet(sheet_name: str) -> bool:
    return _normalize_sheet_tab_name(sheet_name) == "DHF"


def _sheet_supports_multi_project_scope(sheet_name: str) -> bool:
    """同编号多行表示跨项目复用时，合并所属项目/注册国家。"""
    return _is_dhf_sheet(sheet_name)


def _excel_cell_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "1" if value else "0"
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        text = str(value).strip()
        if text.endswith(".0"):
            return text[:-2]
        return text
    return str(value).strip()


def _normalize_excel_header(header: str) -> str:
    text = (header or "").strip().lower()
    return re.sub(r"[\s_\-（）()]+", "", text)


def _lookup_excel_header_field(header: str, header_map: dict[str, str]) -> Optional[str]:
    raw = (header or "").strip()
    if not raw:
        return None
    if raw in header_map:
        return header_map[raw]
    lower = raw.lower()
    lower_map = {k.lower(): v for k, v in header_map.items()}
    if lower in lower_map:
        return lower_map[lower]
    compact = _normalize_excel_header(raw)
    if not compact:
        return None
    for key, field in header_map.items():
        if _normalize_excel_header(key) == compact:
            return field
    return None


def _excel_row_optional_text(row: dict[str, Any], field: str) -> str:
    """读取 Excel 行可选列：列不存在或单元格为空时返回空串，增量导入时不覆盖已有值。"""
    if field not in row:
        return ""
    return (row.get(field) or "").strip()


def _apply_project_code_from_row(
    doc: ControlledDocument, row: dict[str, Any]
) -> bool:
    """有项目编号时才写入；空值不覆盖台账已有 project_code（兼容无该列的 Sheet）。"""
    project_code = _excel_row_optional_text(row, "projectCode")
    if not project_code:
        return False
    if project_code != (doc.project_code or "").strip():
        doc.project_code = project_code
        return True
    return False


def _merge_project_scope_from_row(
    doc: ControlledDocument,
    meta: dict[str, Any],
    row: dict[str, Any],
) -> tuple[dict[str, Any], list[str]]:
    project_name = _excel_row_optional_text(row, "projectName")
    registered_country = _excel_row_optional_text(row, "registeredCountry")
    project_code = _excel_row_optional_text(row, "projectCode")
    updated_fields: list[str] = []
    if not project_name and not registered_country and not project_code:
        return meta, updated_fields
    sheet_category = (
        (row.get("sheetName") or row.get("sheetCategory") or doc.sheet_category or "").strip()
    )
    use_scope_store = (
        _sheet_supports_multi_project_scope(sheet_category)
        or bool(_registration_scope_entries(meta))
        or bool(doc.registration_submitted)
    )
    if not use_scope_store:
        if project_name != _format_scope_display(doc.project_name or ""):
            doc.project_name = _normalize_scope_field_value(project_name)
            if project_name:
                updated_fields.append("所属项目")
        if registered_country != _format_scope_display(doc.registered_country or ""):
            doc.registered_country = _normalize_scope_field_value(registered_country)
            if registered_country:
                updated_fields.append("注册国家")
        if _apply_project_code_from_row(doc, row):
            updated_fields.append("项目编号")
        return meta, updated_fields
    meta = _capture_primary_scope_to_meta(doc, meta)
    before = {
        _scope_pair_key(e.get("projectName", ""), e.get("registeredCountry", ""))
        for e in _registration_scope_entries(meta)
    }
    meta = _add_registration_scope(
        meta,
        project_name=project_name,
        registered_country=registered_country,
        row_index=_excel_row_index_from_row(row),
        project_code=project_code,
    )
    after = {
        _scope_pair_key(e.get("projectName", ""), e.get("registeredCountry", ""))
        for e in _registration_scope_entries(meta)
    }
    prev_code = (doc.project_code or "").strip()
    _apply_registration_scope_display(doc, meta)
    if after != before:
        if project_name:
            updated_fields.append("所属项目")
        if registered_country:
            updated_fields.append("注册国家")
    if project_code and (doc.project_code or "").strip() != prev_code:
        updated_fields.append("项目编号")
    return meta, updated_fields


def _consolidate_sheet_multi_project_scopes(
    preview: list[dict[str, Any]],
    existing: dict[str, ControlledDocument],
    *,
    sheet_name: str,
) -> None:
    if not _sheet_supports_multi_project_scope(sheet_name):
        return
    scopes_by_norm: dict[str, list[dict[str, Any]]] = {}
    for row in preview:
        if (row.get("sheetName") or "").strip() != sheet_name:
            continue
        if row.get("status") in ("not_controlled", "invalid", "duplicate_in_file"):
            continue
        norm = (row.get("normalizedDocumentNumber") or "").strip()
        if not norm:
            norm = normalize_document_number((row.get("documentNumber") or "").strip())
        if not norm:
            continue
        project_name = _excel_row_optional_text(row, "projectName")
        registered_country = _excel_row_optional_text(row, "registeredCountry")
        project_code = _excel_row_optional_text(row, "projectCode")
        if not project_name and not registered_country and not project_code:
            continue
        scopes_by_norm.setdefault(norm, []).append(
            {
                "projectName": project_name,
                "registeredCountry": registered_country,
                "projectCode": project_code,
                "rowIndex": _excel_row_index_from_row(row),
            }
        )
    for norm, scopes in scopes_by_norm.items():
        doc = existing.get(norm)
        if not doc:
            continue
        meta = dict(doc.metadata_json or {}) if isinstance(doc.metadata_json, dict) else {}
        meta = _capture_primary_scope_to_meta(doc, meta)
        before = {
            _scope_pair_key(e.get("projectName", ""), e.get("registeredCountry", ""))
            for e in _registration_scope_entries(meta)
        }
        for scope in scopes:
            meta = _add_registration_scope(
                meta,
                project_name=scope.get("projectName", ""),
                registered_country=scope.get("registeredCountry", ""),
                row_index=scope.get("rowIndex"),
                project_code=scope.get("projectCode", ""),
            )
        after = {
            _scope_pair_key(e.get("projectName", ""), e.get("registeredCountry", ""))
            for e in _registration_scope_entries(meta)
        }
        if after != before:
            _apply_registration_scope_display(doc, meta)
            doc.metadata_json = meta
            doc.updated_at = now_local()
        elif any((scope.get("projectCode") or "").strip() for scope in scopes):
            prev_code = (doc.project_code or "").strip()
            _apply_registration_scope_display(doc, meta)
            if (doc.project_code or "").strip() != prev_code:
                doc.metadata_json = meta
                doc.updated_at = now_local()


def _seen_in_file_entry(
    *,
    preview_index: int,
    priority: int,
    sheet_name: str,
    row: dict[str, Any],
) -> dict[str, Any]:
    return {
        "previewIndex": preview_index,
        "priority": priority,
        "sheetName": sheet_name,
        "rowIndex": int(row.get("rowIndex") or 0),
        "projectName": (row.get("projectName") or "").strip(),
        "registeredCountry": (row.get("registeredCountry") or "").strip(),
    }


def _should_track_registration_link_source(sheet_name: str) -> bool:
    name = (sheet_name or "").strip()
    if not name:
        return False
    if _is_registration_sheet(name) or _is_catalog_sheet(name):
        return False
    return True


def _can_registration_link_source(sheet_name: str) -> bool:
    return _should_track_registration_link_source(sheet_name)


def _load_sheet_order_map() -> dict[str, list[str]]:
    from webapp.models import AppConfig

    row = AppConfig.query.filter_by(config_key=_SHEET_ORDER_CONFIG_KEY).first()
    raw = (row.config_value if row and row.config_value is not None else "") or ""
    if not raw.strip():
        return {}
    try:
        data = json.loads(raw)
    except json.JSONDecodeError:
        return {}
    if not isinstance(data, dict):
        return {}
    out: dict[str, list[str]] = {}
    for org_id, names in data.items():
        if isinstance(names, list):
            out[str(org_id)] = [str(x) for x in names if str(x).strip()]
    return out


def _save_org_sheet_order(org_id: str, sheet_order: list[str]) -> None:
    from webapp.models import AppConfig

    data = _load_sheet_order_map()
    cleaned = [str(x).strip() for x in sheet_order if str(x).strip()]
    data[org_id] = cleaned
    payload = json.dumps(data, ensure_ascii=False)
    row = AppConfig.query.filter_by(config_key=_SHEET_ORDER_CONFIG_KEY).first()
    if row:
        row.config_value = payload
        db.session.add(row)
    else:
        db.session.add(AppConfig(config_key=_SHEET_ORDER_CONFIG_KEY, config_value=payload))


def _apply_sheet_category_display_order(categories: list[str]) -> list[str]:
    """四级表单紧跟程序文件之后展示。"""
    if "四级表单" not in categories or "程序文件" not in categories:
        return categories
    without = [name for name in categories if name != "四级表单"]
    try:
        idx = without.index("程序文件") + 1
    except ValueError:
        return categories
    return without[:idx] + ["四级表单"] + without[idx:]


def _sort_sheet_categories(categories: list[str], org_id: str) -> list[str]:
    """按 Excel 导入 Sheet 顺序排列；无配置时保留调用方顺序（不用名称字母序做 tie-break）。"""
    order = _load_sheet_order_map().get(org_id) or []
    if not order:
        return _apply_sheet_category_display_order(list(categories))
    rank = {name: idx for idx, name in enumerate(order)}
    known = [name for name in order if name in categories]
    unknown = [name for name in categories if name not in rank]
    return _apply_sheet_category_display_order(known + unknown)


def _note_link_source_norm(
    norm: str, sheet_name: str, link_source_norms: dict[str, str]
) -> None:
    if not norm or not _should_track_registration_link_source(sheet_name):
        return
    if norm not in link_source_norms:
        link_source_norms[norm] = (sheet_name or "").strip()


def _registration_target_hit(
    *,
    norm: str,
    document_number: str,
    sheet_name: str,
) -> dict[str, str]:
    return {
        "norm": norm,
        "documentNumber": document_number,
        "sheetName": sheet_name,
    }


def _build_registration_file_indexes(
    rows: list[dict[str, Any]],
) -> tuple[dict[str, dict[str, str]], dict[str, list[dict[str, str]]]]:
    by_norm: dict[str, dict[str, str]] = {}
    by_compare_key: dict[str, list[dict[str, str]]] = {}
    for row in rows:
        sheet_name = (row.get("sheetName") or "").strip()
        if not _should_track_registration_link_source(sheet_name):
            continue
        doc_num = (row.get("documentNumber") or "").strip()
        norm = normalize_document_number(doc_num)
        if not norm or _is_invalid_document_number(doc_num, norm):
            continue
        hit = _registration_target_hit(
            norm=norm,
            document_number=doc_num,
            sheet_name=sheet_name,
        )
        by_norm.setdefault(norm, hit)
        compare_key = registration_compare_key(doc_num)
        if compare_key:
            bucket = by_compare_key.setdefault(compare_key, [])
            if not any(x.get("norm") == norm for x in bucket):
                bucket.append(hit)
    return by_norm, by_compare_key


def _add_registration_target(
    targets: list[dict[str, str]],
    seen_norms: set[str],
    hit: dict[str, str],
    match_type: str,
) -> None:
    norm = (hit.get("norm") or "").strip()
    if not norm or norm in seen_norms:
        return
    seen_norms.add(norm)
    targets.append({**hit, "matchType": match_type})


def _add_registration_target_from_doc(
    targets: list[dict[str, str]],
    seen_norms: set[str],
    doc: ControlledDocument,
    match_type: str,
) -> None:
    if (doc.status or "").strip().lower() == _DOC_STATUS_VOIDED:
        return
    if not _can_registration_link_source(doc.sheet_category or ""):
        return
    norm = (doc.normalized_document_number or "").strip()
    if not norm:
        return
    _add_registration_target(
        targets,
        seen_norms,
        _registration_target_hit(
            norm=norm,
            document_number=(doc.document_number or "").strip(),
            sheet_name=(doc.sheet_category or "").strip(),
        ),
        match_type,
    )


def _resolve_registration_targets(
    reg_row: dict[str, Any],
    *,
    file_by_norm: dict[str, dict[str, str]],
    file_by_compare_key: dict[str, list[dict[str, str]]],
    existing: dict[str, ControlledDocument],
) -> list[dict[str, str]]:
    doc_num = (reg_row.get("documentNumber") or "").strip()
    norm = normalize_document_number(doc_num)
    compare_key = registration_compare_key(doc_num) if doc_num else ""
    targets: list[dict[str, str]] = []
    seen_norms: set[str] = set()

    if norm and norm in file_by_norm:
        _add_registration_target(targets, seen_norms, file_by_norm[norm], "norm")
    if compare_key:
        for hit in file_by_compare_key.get(compare_key, []):
            _add_registration_target(targets, seen_norms, hit, "norm")
    if norm and norm in existing:
        _add_registration_target_from_doc(targets, seen_norms, existing[norm], "norm")
    if compare_key:
        for doc in existing.values():
            if registration_compare_key(doc.document_number or "") != compare_key:
                continue
            _add_registration_target_from_doc(targets, seen_norms, doc, "norm")

    return targets


def _registration_targets_summary(targets: list[dict[str, str]]) -> str:
    if not targets:
        return ""
    parts: list[str] = []
    by_sheet: dict[str, int] = {}
    for item in targets:
        sheet = (item.get("sheetName") or "").strip() or "未知分类"
        by_sheet[sheet] = by_sheet.get(sheet, 0) + 1
    for sheet in sorted(by_sheet):
        count = by_sheet[sheet]
        parts.append(f"{sheet} {count} 条" if count > 1 else sheet)
    return "、".join(parts)


def _excel_header_tokens(field: str) -> frozenset[str]:
    return frozenset(
        k for k, v in _excel_import_header_map().items() if v == field
    )


def _is_excel_noise_row(item: dict[str, Any]) -> bool:
    doc = (item.get("documentNumber") or "").strip()
    title = (item.get("title") or "").strip()
    lifecycle = (item.get("lifecycleStatus") or "").strip()
    if doc in _excel_header_tokens("documentNumber"):
        return True
    if title in _excel_header_tokens("title"):
        return True
    if lifecycle in _excel_header_tokens("lifecycleStatus") and not doc:
        return True
    if not doc and not title and not (item.get("version") or "").strip():
        return True
    return False


def _append_import_log(
    *,
    org_id: str,
    batch_id: str,
    user_id: Optional[str],
    event_type: str,
    row: dict[str, Any],
    reason: str,
    document_id: Optional[str] = None,
) -> None:
    doc_num = (row.get("documentNumber") or "").strip()
    db.session.add(
        DocumentControlImportLog(
            organization_id=org_id,
            import_batch_id=batch_id,
            event_type=event_type,
            document_number=doc_num or None,
            normalized_document_number=normalize_document_number(doc_num) or None,
            sheet_name=(row.get("sheetName") or row.get("sheetCategory") or "").strip() or None,
            row_index=int(row.get("rowIndex") or 0) or None,
            reason=(reason or "")[:512] or None,
            row_payload_json={
                k: row.get(k)
                for k in (
                    "title",
                    "titleEn",
                    "version",
                    "projectName",
                    "registeredCountry",
                    "lifecycleStatus",
                    "status",
                    "statusDetail",
                )
                if row.get(k)
            },
            controlled_document_id=document_id,
            created_by_user_id=user_id,
        )
    )


def _headers_have_status_column(headers: list[str]) -> bool:
    status_keys = {
        k.lower() for k, v in _excel_import_header_map().items() if v == "lifecycleStatus"
    }
    return bool(status_keys & {h.lower() for h in headers if h})


def _is_excel_lifecycle_importable(
    raw: str, *, has_status_column: bool, has_document_number: bool = False
) -> tuple[bool, str]:
    text = (raw or "").strip()
    if not has_status_column:
        return True, ""
    if not text:
        if has_document_number:
            return True, ""
        return False, "状态为空，仅导入受控文件"
    compact = text.replace(" ", "").lower()
    if any(marker in text for marker in _EXCEL_VOIDED_STATUS_MARKERS):
        return False, "状态为作废，不导入"
    if compact in {"void", "voided", "obsolete", "cancelled", "canceled", "inactive"}:
        return False, "状态为作废，不导入"
    if any(marker in text for marker in ("非受控", "不受控")):
        return False, f"状态为「{text}」，非受控"
    if text in _EXCEL_CONTROLLED_STATUS_VALUES or compact in _EXCEL_CONTROLLED_STATUS_VALUES:
        return True, ""
    if text == "受控":
        return True, ""
    return False, f"状态为「{text}」，非受控"


def _sheet_import_priority(sheet_name: str) -> int:
    if _is_catalog_sheet(sheet_name):
        return 1
    return 10


def _is_invalid_document_number(doc_num: str, norm: str) -> bool:
    raw = (doc_num or "").strip()
    if not norm:
        return True
    if raw in {"/", "-", "—", "无", "N/A", "NA", "n/a"}:
        return True
    return False


def _apply_existing_doc_status(
    status: str,
    status_detail: str,
    *,
    norm: str,
    title: str,
    existing: dict[str, ControlledDocument],
) -> tuple[str, str]:
    exists = existing.get(norm)
    if not exists:
        return status, status_detail
    return (
        "update",
        f"台账已有受控记录，将按 Excel 增量更新（编号：{exists.document_number or norm}）",
    )


def _excel_update_preserves_authoritative_names(doc: ControlledDocument, row: dict[str, Any]) -> bool:
    """台账归属程序文件/四级表单时，其他 Sheet 的增量导入不覆盖中文名/英文名。"""
    doc_cat = (doc.sheet_category or "").strip()
    if doc_cat not in _AUTHORITATIVE_NAME_SHEETS:
        return False
    row_sheet = (row.get("sheetName") or row.get("sheetCategory") or "").strip()
    return bool(row_sheet) and row_sheet != doc_cat


def _apply_excel_row_to_document(
    doc: ControlledDocument, row: dict[str, Any]
) -> list[str]:
    updated_fields: list[str] = []
    preserve_names = _excel_update_preserves_authoritative_names(doc, row)
    title = (row.get("title") or "").strip()
    if not preserve_names and title and (doc.title or "").strip() != title:
        doc.title = title
        updated_fields.append("文件名称")
    version = (row.get("version") or "").strip()
    if version != (doc.version or "").strip():
        doc.version = version or None
        if version:
            updated_fields.append("版本")
    title_en = (row.get("titleEn") or "").strip()
    if not preserve_names and title_en and title_en != (doc.title_en or "").strip():
        doc.title_en = title_en
        updated_fields.append("英文名")
    meta = dict(doc.metadata_json or {}) if isinstance(doc.metadata_json, dict) else {}
    meta, scope_updates = _merge_project_scope_from_row(doc, meta, row)
    updated_fields.extend(scope_updates)
    doc_type = (row.get("docTypeCode") or "").strip()
    if doc_type != (doc.doc_type_code or "").strip():
        doc.doc_type_code = doc_type or None
        if doc_type:
            updated_fields.append("文件类型")
    row_index = _excel_row_index_from_row(row)
    if row_index is not None:
        doc.excel_row_index = row_index
    sheet_category = (row.get("sheetName") or row.get("sheetCategory") or "").strip()
    if _is_registration_sheet(sheet_category):
        _set_registration_excel_row_index(doc, row_index)
    batch_id = (row.get("importBatchId") or "").strip()
    if title_en and not preserve_names:
        meta["titleEn"] = title_en
    doc.metadata_json = meta or None
    doc.updated_at = now_local()
    return updated_fields


def _find_excel_header_row(rows: list[tuple]) -> tuple[int, list[str]] | None:
    header_map = _excel_import_header_map()
    doc_keys = {k.lower() for k, v in header_map.items() if v == "documentNumber"}
    for idx, row in enumerate(rows):
        headers = [str(x).strip() if x is not None else "" for x in row]
        normalized_headers = {h.lower() for h in headers if h}
        if doc_keys & normalized_headers:
            return idx, headers
    return None


def _parse_excel_row(headers: list[str], values: list[Any]) -> dict[str, Any]:
    header_map = _excel_import_header_map()
    data: dict[str, str] = {}
    present_fields: set[str] = set()
    for header, value in zip(headers, values):
        key = _lookup_excel_header_field(header, header_map)
        if key:
            data[key] = _excel_cell_text(value)
            present_fields.add(key)
    row: dict[str, Any] = {
        "documentNumber": data.get("documentNumber", ""),
        "title": data.get("title", ""),
        "titleEn": data.get("titleEn", ""),
        "version": data.get("version", ""),
        "docTypeCode": data.get("docTypeCode", ""),
        "projectName": data.get("projectName", ""),
        "registeredCountry": data.get("registeredCountry", ""),
        "lifecycleStatus": data.get("lifecycleStatus", ""),
    }
    # 项目编号为各 Sheet（SOP/DHF/注册文件等）可选列：无表头时不写入键，增量导入不覆盖已有值
    if "projectCode" in present_fields:
        row["projectCode"] = data.get("projectCode", "")
    row["_excelPresentFields"] = sorted(present_fields)
    return row


def _read_excel_rows(
    file_bytes: bytes,
) -> tuple[list[dict[str, Any]], Optional[str], bool, list[str]]:
    wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    all_rows: list[dict[str, Any]] = []
    sheet_order: list[str] = []
    has_status_column = False
    parse_error: Optional[str] = None
    try:
        for ws in wb.worksheets:
            if _is_catalog_sheet(ws.title):
                continue
            raw_rows = list(ws.iter_rows(values_only=True))
            if not raw_rows:
                continue
            found = _find_excel_header_row(raw_rows)
            if found is None:
                continue
            sheet_order.append(ws.title)
            header_idx, headers = found
            sheet_has_status = _headers_have_status_column(headers)
            if sheet_has_status:
                has_status_column = True
            for idx, row in enumerate(raw_rows[header_idx + 1 :], start=header_idx + 2):
                values = [_excel_cell_text(x) for x in row]
                if not any(values):
                    continue
                item = _parse_excel_row(headers, values)
                if _is_excel_noise_row(item):
                    continue
                item["rowIndex"] = idx
                item["sheetName"] = ws.title
                item["hasStatusColumn"] = sheet_has_status
                all_rows.append(item)
        if not all_rows:
            parse_error = (
                "未识别到有效表头，请确认至少一个工作表含「文件编号/编号」等列"
                "（支持：文件名称、版本号、状态、所属项目、注册国家；"
                "项目编号为可选列，SOP/DHF/注册文件等 Sheet 有则导入、无则跳过）"
            )
    finally:
        wb.close()
    return all_rows, parse_error, has_status_column, sheet_order


def _load_existing_docs_by_norm(org_id: str) -> dict[str, ControlledDocument]:
    """导入/判重时仅与受控台账比对；作废编号允许重复。"""
    return load_controlled_docs_by_norm(org_id)


def _build_excel_import_preview(
    rows: list[dict[str, Any]], org_id: str, *, has_status_column: bool = False
) -> list[dict[str, Any]]:
    existing = _load_existing_docs_by_norm(org_id)
    file_by_norm, file_by_compare_key = _build_registration_file_indexes(rows)
    seen_in_file: dict[str, dict[str, Any]] = {}
    reg_targets_by_norm: dict[str, list[dict[str, str]]] = {}
    link_source_norms: dict[str, str] = {}
    preview: list[dict[str, Any]] = []

    def _append_preview(row: dict[str, Any], status: str, status_detail: str) -> int:
        preview.append(
            {
                **row,
                "sheetCategory": row.get("sheetName") or "",
                "lifecycleStatus": (row.get("lifecycleStatus") or "").strip(),
                "normalizedDocumentNumber": normalize_document_number(
                    (row.get("documentNumber") or "").strip()
                ),
                "status": status,
                "statusDetail": status_detail,
            }
        )
        return len(preview) - 1

    for row in rows:
        doc_num = (row.get("documentNumber") or "").strip()
        norm = normalize_document_number(doc_num)
        title = (row.get("title") or "").strip()
        lifecycle_status = (row.get("lifecycleStatus") or "").strip()
        row_has_status_column = bool(row.get("hasStatusColumn"))
        sheet_name = row.get("sheetName") or ""
        priority = _sheet_import_priority(sheet_name)
        if (
            norm
            and not _is_invalid_document_number(doc_num, norm)
            and _should_track_registration_link_source(sheet_name)
        ):
            _note_link_source_norm(norm, sheet_name, link_source_norms)
        importable, lifecycle_detail = _is_excel_lifecycle_importable(
            lifecycle_status,
            has_status_column=row_has_status_column,
            has_document_number=bool(norm) and not _is_invalid_document_number(doc_num, norm),
        )
        status = "new"
        status_detail = ""
        if not importable:
            status = "not_controlled"
            status_detail = lifecycle_detail
        elif _is_invalid_document_number(doc_num, norm):
            status = "invalid"
            status_detail = "缺少有效文件编号"
        elif _is_registration_sheet(sheet_name):
            reg_targets = _resolve_registration_targets(
                row,
                file_by_norm=file_by_norm,
                file_by_compare_key=file_by_compare_key,
                existing=existing,
            )
            project_name = _excel_row_optional_text(row, "projectName")
            registered_country = _excel_row_optional_text(row, "registeredCountry")
            project_code = _excel_row_optional_text(row, "projectCode")
            if not reg_targets and norm in reg_targets_by_norm and (
                project_name or registered_country or project_code
            ):
                reg_targets = list(reg_targets_by_norm[norm])
                row["_registrationTargets"] = reg_targets
                status = "registration_update"
                status_detail = "将补充注册所属项目/国家/项目编号至已关联台账"
            elif reg_targets:
                status = "registration_update"
                row["_registrationTargets"] = reg_targets
                reg_targets_by_norm[norm] = reg_targets
                summary = _registration_targets_summary(reg_targets)
                status_detail = (
                    f"将关联注册文件清单至：{summary}" if summary else "将关联注册文件清单"
                )
            elif norm in seen_in_file:
                prev = seen_in_file[norm]
                status = "duplicate_in_file"
                status_detail = (
                    f"与 {prev.get('sheetName')}!{prev.get('rowIndex')} 行重复"
                )
            else:
                status, status_detail = _apply_existing_doc_status(
                    "new",
                    "",
                    norm=norm,
                    title=title,
                    existing=existing,
                )
        elif norm in seen_in_file:
            prev = seen_in_file[norm]
            project_name = _excel_row_optional_text(row, "projectName")
            registered_country = _excel_row_optional_text(row, "registeredCountry")
            curr_pair = _scope_pair_key(project_name, registered_country)
            prev_pair = _scope_pair_key(
                (prev.get("projectName") or "").strip(),
                (prev.get("registeredCountry") or "").strip(),
            )
            if (
                prev.get("sheetName") == sheet_name
                and _sheet_supports_multi_project_scope(sheet_name)
                and curr_pair != ("", "")
                and curr_pair != prev_pair
            ):
                status, status_detail = _apply_existing_doc_status(
                    "update",
                    (
                        f"同编号将补充所属项目/国家"
                        f"（{sheet_name} 第 {int(row.get('rowIndex') or 0)} 行）"
                    ),
                    norm=norm,
                    title=title,
                    existing=existing,
                )
                preview_index = _append_preview(row, status, status_detail)
                seen_in_file[norm] = _seen_in_file_entry(
                    preview_index=preview_index,
                    priority=priority,
                    sheet_name=sheet_name,
                    row=row,
                )
                _note_link_source_norm(norm, sheet_name, link_source_norms)
                continue
            if priority > int(prev.get("priority") or 0):
                preview[prev["previewIndex"]]["status"] = "duplicate_in_file"
                preview[prev["previewIndex"]]["statusDetail"] = (
                    f"与 {sheet_name} 中同编号记录重复，已保留 {sheet_name} 分类"
                )
                status, status_detail = _apply_existing_doc_status(
                    "new",
                    "",
                    norm=norm,
                    title=title,
                    existing=existing,
                )
                seen_in_file[norm] = {
                    "previewIndex": _append_preview(row, status, status_detail),
                    "priority": priority,
                    "sheetName": sheet_name,
                    "rowIndex": int(row.get("rowIndex") or 0),
                }
                _note_link_source_norm(norm, sheet_name, link_source_norms)
                continue
            status = "duplicate_in_file"
            status_detail = (
                f"与 {prev.get('sheetName')}!{prev.get('rowIndex')} 行重复"
                if prev.get("sheetName")
                else f"与第 {prev.get('rowIndex')} 行重复"
            )
        else:
            status, status_detail = _apply_existing_doc_status(
                "new",
                "",
                norm=norm,
                title=title,
                existing=existing,
            )
            preview_index = _append_preview(row, status, status_detail)
            if status in ("new", "update"):
                seen_in_file[norm] = _seen_in_file_entry(
                    preview_index=preview_index,
                    priority=priority,
                    sheet_name=sheet_name,
                    row=row,
                )
            _note_link_source_norm(norm, sheet_name, link_source_norms)
            continue

        preview_index = _append_preview(row, status, status_detail)
        if status in ("new", "update"):
            seen_in_file[norm] = _seen_in_file_entry(
                preview_index=preview_index,
                priority=priority,
                sheet_name=sheet_name,
                row=row,
            )
        elif status == "registration_update" and row.get("_registrationTargets"):
            reg_targets_by_norm[norm] = list(row.get("_registrationTargets") or [])
        _note_link_source_norm(norm, sheet_name, link_source_norms)

    return preview


def _apply_registration_update(
    doc: ControlledDocument, row: dict[str, Any]
) -> None:
    doc.registration_submitted = True
    project_name = _excel_row_optional_text(row, "projectName")
    registered_country = _excel_row_optional_text(row, "registeredCountry")
    project_code = _excel_row_optional_text(row, "projectCode")
    title_en = (row.get("titleEn") or "").strip()
    row_index = _excel_row_index_from_row(row)
    updated_fields: list[str] = []
    if title_en:
        doc.title_en = title_en
        updated_fields.append("英文名")
    meta = dict(doc.metadata_json or {}) if isinstance(doc.metadata_json, dict) else {}
    meta = _capture_primary_scope_to_meta(doc, meta)
    if project_name or registered_country or project_code:
        meta = _add_registration_scope(
            meta,
            project_name=project_name,
            registered_country=registered_country,
            row_index=row_index,
            project_code=project_code,
        )
        prev_code = (doc.project_code or "").strip()
        _apply_registration_scope_display(doc, meta)
        if project_name:
            updated_fields.append("所属项目")
        if registered_country:
            updated_fields.append("注册国家")
        if project_code and (doc.project_code or "").strip() != prev_code:
            updated_fields.append("项目编号")
    _set_registration_excel_row_index(doc, row_index)
    if title_en:
        meta["titleEn"] = title_en
    meta["registrationLinkedFromSheet"] = _REGISTRATION_SHEET_NAME
    meta["registrationLinkedAt"] = now_local().isoformat()
    doc.metadata_json = meta
    doc.updated_at = now_local()
    row["_registrationUpdatedFields"] = updated_fields


def _import_project_code_block_reason(
    org_id: str,
    project_code: str,
    *,
    project_id: Optional[str] = None,
    project_name: Optional[str] = None,
    registered_country: Optional[str] = None,
    exclude_document_id: Optional[str] = None,
    confirm_sync: bool = False,
) -> Optional[str]:
    gate = gate_project_code_save(
        org_id,
        project_code,
        project_id=project_id,
        project_name=project_name,
        registered_country=registered_country,
        exclude_document_id=exclude_document_id,
        confirm_sync=confirm_sync,
    )
    if not gate:
        return None
    if gate[0] == "confirm":
        payload = gate[1] if isinstance(gate[1], dict) else {}
        return (payload.get("message") or "项目编号变更需确认") + (
            "；请先在页面确认修改后再导入，或导入时勾选同步确认"
        )
    return str(gate[1])


def _import_excel_rows(
    preview: list[dict[str, Any]],
    org_id: str,
    batch_id: str,
    user_id: Optional[str],
    *,
    confirm_project_code_sync: bool = False,
) -> tuple[int, int, int, list[dict[str, Any]]]:
    existing = _load_existing_docs_by_norm(org_id)
    imported_in_batch: set[str] = set()
    updated_in_batch: set[str] = set()
    imported = 0
    updated = 0
    registration_updated = 0
    skipped: list[dict[str, Any]] = []

    for row in preview:
        if row.get("status") != "new":
            continue
        doc_num = (row.get("documentNumber") or "").strip()
        norm = normalize_document_number(doc_num)
        if not norm:
            reason = "缺少文件编号"
            skipped.append({**row, "skipReason": reason})
            _append_import_log(
                org_id=org_id,
                batch_id=batch_id,
                user_id=user_id,
                event_type="import_skip",
                row=row,
                reason=reason,
            )
            continue
        if norm in imported_in_batch:
            reason = row.get("statusDetail") or "本批导入中受控编号重复"
            skipped.append({**row, "skipReason": reason})
            _append_import_log(
                org_id=org_id,
                batch_id=batch_id,
                user_id=user_id,
                event_type="import_skip",
                row=row,
                reason=reason,
            )
            continue
        conflict_msg = controlled_number_conflict_message(org_id, norm)
        if conflict_msg or norm in existing:
            reason = row.get("statusDetail") or conflict_msg or "受控编号已存在"
            skipped.append({**row, "skipReason": reason})
            _append_import_log(
                org_id=org_id,
                batch_id=batch_id,
                user_id=user_id,
                event_type="import_skip",
                row=row,
                reason=reason,
            )
            continue
        title_en = (row.get("titleEn") or "").strip()
        row_index = _excel_row_index_from_row(row)
        sheet_cat = (row.get("sheetName") or row.get("sheetCategory") or "").strip() or None
        project_name = _excel_row_optional_text(row, "projectName")
        registered_country = _excel_row_optional_text(row, "registeredCountry")
        project_code = _excel_row_optional_text(row, "projectCode")
        metadata_json: Optional[dict[str, Any]] = {"titleEn": title_en} if title_en else None
        display_project = _normalize_scope_field_value(project_name)
        display_country = _normalize_scope_field_value(registered_country)
        display_code = _normalize_scope_field_value(project_code)
        if display_code:
            pcode_conflict = _import_project_code_block_reason(
                org_id,
                display_code,
                project_name=display_project or project_name,
                registered_country=display_country or registered_country,
                confirm_sync=confirm_project_code_sync,
            )
            if pcode_conflict:
                skipped.append({**row, "skipReason": pcode_conflict})
                _append_import_log(
                    org_id=org_id,
                    batch_id=batch_id,
                    user_id=user_id,
                    event_type="import_skip",
                    row=row,
                    reason=pcode_conflict,
                )
                continue
        if _sheet_supports_multi_project_scope(sheet_cat or "") and (
            project_name or registered_country or project_code
        ):
            meta: dict[str, Any] = dict(metadata_json or {})
            meta = _add_registration_scope(
                meta,
                project_name=project_name,
                registered_country=registered_country,
                row_index=row_index,
                project_code=project_code,
            )
            scope_project, scope_country, scope_code = _scope_display_fields_from_meta(meta)
            display_project = scope_project or display_project
            display_country = scope_country or display_country
            display_code = scope_code or display_code
            metadata_json = meta
        doc = ControlledDocument(
            organization_id=org_id,
            document_number=doc_num,
            normalized_document_number=norm,
            version=(row.get("version") or "").strip() or None,
            title=(row.get("title") or "").strip() or "未命名文件",
            title_en=title_en or None,
            doc_type_code=(row.get("docTypeCode") or "").strip() or None,
            project_code=display_code,
            project_name=display_project,
            registered_country=display_country,
            sheet_category=sheet_cat,
            excel_row_index=row_index,
            registration_excel_row_index=row_index if _is_registration_sheet(sheet_cat or "") else None,
            status=_DOC_STATUS_CONTROLLED,
            metadata_json=metadata_json,
            source="excel_import",
            import_batch_id=batch_id,
            created_by_user_id=user_id,
        )
        try:
            with db.session.begin_nested():
                db.session.add(doc)
                db.session.flush()
                link_controlled_document_to_page1_project(doc, organization_id=org_id)
        except IntegrityError:
            reason = "受控编号已存在（并发冲突）"
            skipped.append({**row, "skipReason": reason})
            _append_import_log(
                org_id=org_id,
                batch_id=batch_id,
                user_id=user_id,
                event_type="import_fail",
                row=row,
                reason=reason,
            )
            existing = _load_existing_docs_by_norm(org_id)
            continue
        imported_in_batch.add(norm)
        existing[norm] = doc
        imported += 1
        _append_import_log(
            org_id=org_id,
            batch_id=batch_id,
            user_id=user_id,
            event_type="import_success",
            row=row,
            reason="导入成功",
            document_id=doc.id,
        )

    for row in preview:
        if row.get("status") != "update":
            continue
        doc_num = (row.get("documentNumber") or "").strip()
        norm = normalize_document_number(doc_num)
        if not norm:
            continue
        if norm in updated_in_batch:
            continue
        doc = existing.get(norm)
        if not doc:
            reason = "未找到可更新的受控台账记录"
            skipped.append({**row, "skipReason": reason})
            _append_import_log(
                org_id=org_id,
                batch_id=batch_id,
                user_id=user_id,
                event_type="import_fail",
                row=row,
                reason=reason,
            )
            continue
        row_with_batch = {**row, "importBatchId": batch_id}
        project_code_in_row = _excel_row_optional_text(row, "projectCode")
        if project_code_in_row:
            project_name_in_row = _excel_row_optional_text(row, "projectName")
            country_in_row = _excel_row_optional_text(row, "registeredCountry")
            eff_name = (
                _normalize_scope_field_value(project_name_in_row)
                or doc.project_name
            )
            eff_country = (
                _normalize_scope_field_value(country_in_row)
                or doc.registered_country
            )
            pcode_conflict = _import_project_code_block_reason(
                org_id,
                project_code_in_row,
                project_id=doc.project_id,
                project_name=eff_name,
                registered_country=eff_country,
                exclude_document_id=doc.id,
                confirm_sync=confirm_project_code_sync,
            )
            if pcode_conflict:
                skipped.append({**row, "skipReason": pcode_conflict})
                _append_import_log(
                    org_id=org_id,
                    batch_id=batch_id,
                    user_id=user_id,
                    event_type="import_fail",
                    row=row,
                    reason=pcode_conflict,
                )
                continue
        updated_fields = _apply_excel_row_to_document(doc, row_with_batch)
        link_controlled_document_to_page1_project(doc, organization_id=org_id)
        if not doc.import_batch_id:
            doc.import_batch_id = batch_id
        if norm not in imported_in_batch:
            updated += 1
        updated_in_batch.add(norm)
        reason = "增量更新成功"
        if updated_fields:
            reason = f"{reason}：已更新 {'、'.join(updated_fields)}"
        elif (row.get("statusDetail") or "").strip():
            reason = row.get("statusDetail") or reason
        _append_import_log(
            org_id=org_id,
            batch_id=batch_id,
            user_id=user_id,
            event_type="import_update",
            row=row,
            reason=reason,
            document_id=doc.id,
        )

    for row in preview:
        if row.get("status") != "registration_update":
            continue
        targets = row.get("_registrationTargets") or []
        if not targets:
            doc_num = (row.get("documentNumber") or "").strip()
            norm = normalize_document_number(doc_num)
            if norm:
                targets = [{"norm": norm, "matchType": "norm"}]
        linked = 0
        for target in targets:
            tgt_norm = (target.get("norm") or "").strip()
            if not tgt_norm:
                continue
            doc = existing.get(tgt_norm)
            if not doc:
                continue
            if not _can_registration_link_source(doc.sheet_category or ""):
                continue
            if (doc.status or "").strip().lower() == _DOC_STATUS_VOIDED:
                continue
            project_code_in_row = _excel_row_optional_text(row, "projectCode")
            if project_code_in_row:
                project_name_in_row = _excel_row_optional_text(row, "projectName")
                country_in_row = _excel_row_optional_text(row, "registeredCountry")
                eff_name = (
                    _normalize_scope_field_value(project_name_in_row)
                    or doc.project_name
                )
                eff_country = (
                    _normalize_scope_field_value(country_in_row)
                    or doc.registered_country
                )
                pcode_conflict = _import_project_code_block_reason(
                    org_id,
                    project_code_in_row,
                    project_id=doc.project_id,
                    project_name=eff_name,
                    registered_country=eff_country,
                    exclude_document_id=doc.id,
                    confirm_sync=confirm_project_code_sync,
                )
                if pcode_conflict:
                    skipped.append({**row, "skipReason": pcode_conflict})
                    _append_import_log(
                        org_id=org_id,
                        batch_id=batch_id,
                        user_id=user_id,
                        event_type="import_fail",
                        row=row,
                        reason=pcode_conflict,
                    )
                    continue
            link_row = {
                **row,
                "statusDetail": (
                    f"注册关联（{target.get('matchType') or 'match'}）"
                    f"：{doc.sheet_category or '-'} / {doc.document_number or tgt_norm}"
                ),
            }
            _apply_registration_update(doc, link_row)
            registration_updated += 1
            linked += 1
            updated_fields = link_row.get("_registrationUpdatedFields") or []
            link_reason = link_row.get("statusDetail") or "已关联注册文件清单"
            if updated_fields:
                link_reason = f"{link_reason}；已更新：{'、'.join(updated_fields)}"
            _append_import_log(
                org_id=org_id,
                batch_id=batch_id,
                user_id=user_id,
                event_type="registration_link",
                row={
                    **link_row,
                    "rowIndex": _excel_row_index_from_row(link_row) or link_row.get("rowIndex"),
                    "sheetName": _REGISTRATION_SHEET_NAME,
                },
                reason=link_reason,
                document_id=doc.id,
            )
        if not linked:
            reason = "未找到可关联的受控台账记录（注册文件编号与程序文件/四级表单/SOP/DHF 规范后仍无一致编号）"
            skipped.append({**row, "skipReason": reason})
            _append_import_log(
                org_id=org_id,
                batch_id=batch_id,
                user_id=user_id,
                event_type="import_fail",
                row=row,
                reason=reason,
            )

    # 汇总同一关联目标在注册文件中多行填写的项目/国家，避免漏并
    scopes_by_target: dict[str, list[dict[str, Any]]] = {}
    for row in preview:
        if row.get("status") != "registration_update":
            continue
        project_name = _excel_row_optional_text(row, "projectName")
        registered_country = _excel_row_optional_text(row, "registeredCountry")
        project_code = _excel_row_optional_text(row, "projectCode")
        if not project_name and not registered_country and not project_code:
            continue
        scope = {
            "projectName": project_name,
            "registeredCountry": registered_country,
            "projectCode": project_code,
            "rowIndex": _excel_row_index_from_row(row),
        }
        for target in row.get("_registrationTargets") or []:
            tgt_norm = (target.get("norm") or "").strip()
            if not tgt_norm:
                continue
            scopes_by_target.setdefault(tgt_norm, []).append(scope)
    for tgt_norm, scopes in scopes_by_target.items():
        doc = existing.get(tgt_norm)
        if not doc or not _can_registration_link_source(doc.sheet_category or ""):
            continue
        if (doc.status or "").strip().lower() == _DOC_STATUS_VOIDED:
            continue
        meta = dict(doc.metadata_json or {}) if isinstance(doc.metadata_json, dict) else {}
        meta = _capture_primary_scope_to_meta(doc, meta)
        before = {
            _scope_pair_key(e.get("projectName", ""), e.get("registeredCountry", ""))
            for e in _registration_scope_entries(meta)
        }
        for scope in scopes:
            meta = _add_registration_scope(
                meta,
                project_name=scope.get("projectName", ""),
                registered_country=scope.get("registeredCountry", ""),
                row_index=scope.get("rowIndex"),
                project_code=scope.get("projectCode", ""),
            )
        after = {
            _scope_pair_key(e.get("projectName", ""), e.get("registeredCountry", ""))
            for e in _registration_scope_entries(meta)
        }
        if after != before:
            doc.registration_submitted = True
            _apply_registration_scope_display(doc, meta)
            meta["registrationLinkedFromSheet"] = _REGISTRATION_SHEET_NAME
            meta.setdefault("registrationLinkedAt", now_local().isoformat())
            doc.metadata_json = meta
            doc.updated_at = now_local()
        elif any((scope.get("projectCode") or "").strip() for scope in scopes):
            prev_code = (doc.project_code or "").strip()
            _apply_registration_scope_display(doc, meta)
            if (doc.project_code or "").strip() != prev_code:
                doc.registration_submitted = True
                meta["registrationLinkedFromSheet"] = _REGISTRATION_SHEET_NAME
                meta.setdefault("registrationLinkedAt", now_local().isoformat())
                doc.metadata_json = meta
                doc.updated_at = now_local()

    _consolidate_sheet_multi_project_scopes(preview, existing, sheet_name="DHF")

    skip_statuses = {
        "duplicate_in_file",
        "not_controlled",
        "invalid",
    }
    for row in preview:
        status = row.get("status") or ""
        if status in ("new", "update", "registration_update", "import_success"):
            continue
        if status in skip_statuses:
            _append_import_log(
                org_id=org_id,
                batch_id=batch_id,
                user_id=user_id,
                event_type="import_skip",
                row=row,
                reason=row.get("statusDetail") or status,
            )

    db.session.commit()
    return imported, updated, registration_updated, skipped


@document_control_bp.post("/api/document-control/import/excel")
def api_document_control_import_excel():
    blocked = _require_feature()
    if blocked is not None:
        return blocked
    wall = login_wall()
    if wall is not None:
        return wall
    org_id, _ = _org_context()
    file = request.files.get("file")
    if not file or not file.filename:
        return jsonify({"message": "请上传 Excel 文件"}), 400
    rows, parse_error, has_status_column, sheet_order = _read_excel_rows(file.read())
    if parse_error:
        return jsonify({"message": parse_error}), 400
    preview = _build_excel_import_preview(
        rows, org_id, has_status_column=has_status_column
    )
    confirm = str(request.form.get("confirm") or "").strip().lower() in (
        "1",
        "true",
        "yes",
    )
    confirm_project_code_sync = str(
        request.form.get("confirmProjectCodeSync") or ""
    ).strip().lower() in ("1", "true", "yes", "on")
    imported = 0
    updated = 0
    registration_updated = 0
    skipped: list[dict[str, Any]] = []
    batch_id = uuid.uuid4().hex[:12]
    if confirm:
        user_id = (session.get("user_id") or "").strip() or None
        try:
            imported, updated, registration_updated, skipped = _import_excel_rows(
                preview,
                org_id,
                batch_id,
                user_id,
                confirm_project_code_sync=confirm_project_code_sync or confirm,
            )
            _save_org_sheet_order(org_id, sheet_order)
        except IntegrityError:
            db.session.rollback()
            return jsonify(
                {
                    "message": "导入失败：存在重复编号，请刷新预览后重试",
                    "preview": preview,
                    "confirm": True,
                    "imported": 0,
                    "updated": 0,
                    "registrationUpdated": 0,
                    "skipped": skipped,
                    "batchId": batch_id,
                }
            ), 409
    summary = {
        "total": len(preview),
        "new": sum(1 for x in preview if x.get("status") == "new"),
        "update": sum(1 for x in preview if x.get("status") == "update"),
        "registrationUpdate": sum(
            1 for x in preview if x.get("status") == "registration_update"
        ),
        "duplicateInFile": sum(1 for x in preview if x.get("status") == "duplicate_in_file"),
        "invalid": sum(1 for x in preview if x.get("status") == "invalid"),
        "notControlled": sum(1 for x in preview if x.get("status") == "not_controlled"),
        "hasStatusColumn": has_status_column,
        "sheetOrder": sheet_order,
    }
    return jsonify(
        {
            "preview": preview,
            "summary": summary,
            "confirm": confirm,
            "imported": imported,
            "updated": updated,
            "registrationUpdated": registration_updated,
            "skipped": skipped,
            "batchId": batch_id if confirm else None,
        }
    )


@document_control_bp.post("/api/document-control/import/files")
def api_document_control_import_files():
    blocked = _require_feature()
    if blocked is not None:
        return blocked
    wall = login_wall()
    if wall is not None:
        return wall
    org_id, collection = _org_context()
    base = integration_api_base()
    if not base:
        return jsonify({"message": "未配置文档服务地址"}), 400
    files = request.files.getlist("files")
    multipart = [("files", (f.filename, f.stream, f.mimetype)) for f in files if f and f.filename]
    if not multipart:
        return jsonify({"message": "请上传文件"}), 400
    try:
        resp = requests.post(
            f"{base}/api/integration/document-control/extract-batch",
            data={"collection": collection},
            files=multipart,
            headers=upstream_headers(for_multipart=True, organization_id=org_id),
            timeout=integration_requests_timeout(read_seconds=180),
        )
    except requests.RequestException as exc:
        return jsonify({"message": format_upstream_request_error(exc, base)}), 502
    if resp.status_code != 200:
        return jsonify(
            {
                "message": user_facing_upstream_error(
                    f"上游提取失败 HTTP {resp.status_code}",
                    "文档识别服务暂不可用，请稍后重试",
                )
            }
        ), 502
    data = resp.json()
    return jsonify(data)


@document_control_bp.post("/api/document-control/schemes/sync-from-kb")
def api_document_control_sync_from_kb():
    blocked = _require_feature()
    if blocked is not None:
        return blocked
    wall = login_wall()
    if wall is not None:
        return wall
    org_id, collection = _org_context()
    base = integration_api_base()
    if not base:
        return jsonify({"message": "未配置文档服务地址"}), 400
    body = request.get_json(force=True) or {}
    payload = {
        "collection": collection,
        "query": (body.get("query") or "文件控制程序 编号规则").strip(),
    }
    try:
        resp = requests.post(
            f"{base}/api/integration/document-control/parse-numbering-rules",
            json=payload,
            headers=upstream_headers(for_multipart=False, organization_id=org_id),
            timeout=integration_requests_timeout(read_seconds=120),
        )
    except requests.RequestException as exc:
        return jsonify({"message": format_upstream_request_error(exc, base)}), 502
    if resp.status_code != 200:
        return jsonify({"message": f"规则解析失败 HTTP {resp.status_code}"}), 502
    data = resp.json() if resp.content else {}
    if not isinstance(data, dict):
        data = {}
    rules = data.get("rules") or data.get("candidates") or []
    created, updated, skipped = upsert_schemes_from_kb_rules(org_id, rules)
    refs = data.get("references") or []
    message = (data.get("message") or "").strip()
    if created or updated:
        message = f"{message}；已写入规则表（新增 {created}，更新 {updated}）".strip("；")
    elif rules:
        message = message or f"规则已与库内一致（共 {len(rules)} 类）"
    elif not message:
        message = f"检索到 {len(refs)} 条片段，未解析出可写入的规则"
    rows = (
        NumberingScheme.query.filter_by(organization_id=org_id)
        .order_by(NumberingScheme.created_at.desc())
        .all()
    )
    return jsonify(
        {
            **data,
            "message": message,
            "created": created,
            "updated": updated,
            "skipped": skipped,
            "items": [_serialize_scheme(x) for x in rows],
        }
    )

