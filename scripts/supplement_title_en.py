from __future__ import annotations

"""按文件编号从文控外部资料批量补充 title_en（英文名）。

数据来源（默认）：
- Excel：DHF文件涉及法规清单（CE&FDA）.xlsx — Sheet1「文件编号」「英文名」
- Word：英文目录.docx — 表格单元格内「编号 + 英文标题」

合并规则：同编号时 Excel 优先，Word 仅补 Excel 没有的项。
匹配规则：normalize_document_number 对齐台账 normalized_document_number；
可选 registration_compare_key 二次匹配（去连字符）。

用法：
    # 预览（默认，不写库）
    python scripts/supplement_title_en.py

    # 确认后写入（仅填空 title_en）
    python scripts/supplement_title_en.py --apply

    # 自定义路径与组织
    python scripts/supplement_title_en.py --org-id <uuid> --xlsx <path> --docx <path>
"""

import argparse
import csv
import re
import sys
from dataclasses import dataclass, field
from pathlib import Path
from typing import Literal

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

DEFAULT_ORG_ID = "9145f16e-90ab-4f90-b2b2-0b95bcc9da45"
DEFAULT_XLSX = (
    r"g:\互联网产品部\质量体系\注册体系文件"
    r"\【05】通用规范\国外\DHF文件涉及法规清单（CE&FDA）.xlsx"
)
DEFAULT_DOCX = (
    r"g:\互联网产品部\质量体系\注册体系文件"
    r"\【02】管理体系\【06】其他\英文目录.docx"
)

_SKIP_XLSX_NUMBERS = frozenset({"文件编号", "网络安全相关文件"})
_NUM_TITLE_RE = re.compile(
    r"(?:^|[\s_/])([A-Z]{2,}(?:-[A-Z0-9]+)+)\s+(.+)$",
    re.IGNORECASE,
)

SourceKind = Literal["xlsx", "docx"]


@dataclass
class SourceEntry:
    raw_number: str
    title_en: str
    source: SourceKind


@dataclass
class PreviewRow:
    document_number: str
    title_zh: str
    title_en_new: str
    source: SourceKind
    doc_id: str


@dataclass
class SkipRow:
    document_number: str
    title_zh: str
    title_en_existing: str
    title_en_source: str
    source: SourceKind


@dataclass
class ConflictRow:
    document_number: str
    title_zh: str
    title_en_existing: str
    title_en_new: str
    source: SourceKind


@dataclass
class UnmatchedEntry:
    raw_number: str
    title_en: str
    source: SourceKind


@dataclass
class PreviewReport:
    org_id: str
    source_count: int
    ledger_count: int
    to_fill: list[PreviewRow] = field(default_factory=list)
    to_skip: list[SkipRow] = field(default_factory=list)
    conflicts: list[ConflictRow] = field(default_factory=list)
    unmatched: list[UnmatchedEntry] = field(default_factory=list)
    compare_key_matches: int = 0


def _fail(msg: str) -> None:
    print(f"[FAIL] {msg}", file=sys.stderr)
    sys.exit(1)


def parse_excel(xlsx_path: Path) -> dict[str, SourceEntry]:
    from openpyxl import load_workbook

    from webapp.document_control.numbering_engine import normalize_document_number

    if not xlsx_path.is_file():
        _fail(f"Excel 不存在: {xlsx_path}")

    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    try:
        ws = wb["Sheet1"] if "Sheet1" in wb.sheetnames else wb.active
        result: dict[str, SourceEntry] = {}
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0:
                continue
            vals = [str(c).strip() if c is not None else "" for c in row]
            if len(vals) < 4:
                continue
            raw_num, title_en = vals[1], vals[3]
            if not raw_num or not title_en or raw_num in _SKIP_XLSX_NUMBERS:
                continue
            norm = normalize_document_number(raw_num)
            if not norm:
                continue
            result[norm] = SourceEntry(raw_number=raw_num, title_en=title_en.strip(), source="xlsx")
        return result
    finally:
        wb.close()


def parse_docx(docx_path: Path) -> dict[str, SourceEntry]:
    from docx import Document

    from webapp.document_control.numbering_engine import normalize_document_number

    if not docx_path.is_file():
        _fail(f"Word 不存在: {docx_path}")

    result: dict[str, SourceEntry] = {}
    doc = Document(str(docx_path))
    for table in doc.tables:
        for row in table.rows:
            for cell in row.cells:
                text = cell.text.strip().replace("\n", " ")
                if not text:
                    continue
                m = _NUM_TITLE_RE.search(text)
                if not m:
                    continue
                raw_num = m.group(1).strip()
                title_en = m.group(2).strip()
                norm = normalize_document_number(raw_num)
                if not norm or not title_en:
                    continue
                result[norm] = SourceEntry(
                    raw_number=raw_num,
                    title_en=title_en,
                    source="docx",
                )
    return result


def merge_sources(
    xlsx_map: dict[str, SourceEntry],
    docx_map: dict[str, SourceEntry],
) -> dict[str, SourceEntry]:
    merged = dict(docx_map)
    merged.update(xlsx_map)
    return merged


def build_ledger_indexes(org_id: str):
    from webapp.document_control.numbering_engine import registration_compare_key
    from webapp.models import ControlledDocument

    docs = ControlledDocument.query.filter_by(organization_id=org_id).all()
    by_norm: dict[str, ControlledDocument] = {}
    by_compare: dict[str, ControlledDocument] = {}
    for doc in docs:
        norm = (doc.normalized_document_number or "").strip()
        if norm and norm not in by_norm:
            by_norm[norm] = doc
        cmp_key = registration_compare_key(doc.document_number or norm)
        if cmp_key and cmp_key not in by_compare:
            by_compare[cmp_key] = doc
    return docs, by_norm, by_compare


def build_preview(
    org_id: str,
    source_map: dict[str, SourceEntry],
    *,
    use_compare_key: bool = True,
) -> PreviewReport:
    from webapp.document_control.numbering_engine import (
        normalize_document_number,
        registration_compare_key,
    )

    docs, by_norm, by_compare = build_ledger_indexes(org_id)
    report = PreviewReport(
        org_id=org_id,
        source_count=len(source_map),
        ledger_count=len(docs),
    )
    matched_norms: set[str] = set()

    for norm, entry in source_map.items():
        doc = by_norm.get(norm)
        used_compare = False
        if not doc and use_compare_key:
            cmp_key = registration_compare_key(entry.raw_number)
            if cmp_key:
                doc = by_compare.get(cmp_key)
                if doc:
                    used_compare = True
                    report.compare_key_matches += 1
        if not doc:
            report.unmatched.append(
                UnmatchedEntry(
                    raw_number=entry.raw_number,
                    title_en=entry.title_en,
                    source=entry.source,
                )
            )
            continue

        matched_norms.add(norm)
        cur = (doc.title_en or "").strip()
        if not cur:
            report.to_fill.append(
                PreviewRow(
                    document_number=doc.document_number,
                    title_zh=doc.title or "",
                    title_en_new=entry.title_en,
                    source=entry.source,
                    doc_id=doc.id,
                )
            )
        elif cur.lower() == entry.title_en.lower():
            report.to_skip.append(
                SkipRow(
                    document_number=doc.document_number,
                    title_zh=doc.title or "",
                    title_en_existing=cur,
                    title_en_source=entry.title_en,
                    source=entry.source,
                )
            )
        else:
            report.conflicts.append(
                ConflictRow(
                    document_number=doc.document_number,
                    title_zh=doc.title or "",
                    title_en_existing=cur,
                    title_en_new=entry.title_en,
                    source=entry.source,
                )
            )

    return report


def apply_fill(org_id: str, preview: PreviewReport) -> int:
    from webapp import db
    from webapp.models import ControlledDocument, now_local

    updated = 0
    for row in preview.to_fill:
        doc = ControlledDocument.query.filter_by(
            id=row.doc_id,
            organization_id=org_id,
        ).first()
        if not doc:
            continue
        if (doc.title_en or "").strip():
            continue
        doc.title_en = row.title_en_new
        meta = dict(doc.metadata_json or {}) if isinstance(doc.metadata_json, dict) else {}
        meta["titleEn"] = row.title_en_new
        doc.metadata_json = meta
        doc.updated_at = now_local()
        updated += 1
    if updated:
        db.session.commit()
    return updated


def print_summary(preview: PreviewReport, *, applied: int | None = None) -> None:
    matched = len(preview.to_fill) + len(preview.to_skip) + len(preview.conflicts)
    print("=" * 60)
    print("文控台账英文名补充 — 报告")
    print("=" * 60)
    print(f"组织 ID     : {preview.org_id}")
    print(f"外部编号数  : {preview.source_count}")
    print(f"台账记录数  : {preview.ledger_count}")
    print(f"台账可匹配  : {matched}")
    if preview.compare_key_matches:
        print(f"  (其中 compare_key 二次匹配: {preview.compare_key_matches})")
    print(f"将填充      : {len(preview.to_fill)}")
    print(f"已有且一致  : {len(preview.to_skip)}")
    print(f"英文冲突    : {len(preview.conflicts)}")
    print(f"外部无匹配  : {len(preview.unmatched)}")
    if applied is not None:
        print(f"已写入      : {applied}")
    print()

    if preview.to_fill:
        print(f"--- 将填充 ({len(preview.to_fill)}) ---")
        for r in preview.to_fill[:20]:
            print(f"  {r.document_number}\t{r.title_en_new}\t[{r.source}]")
        if len(preview.to_fill) > 20:
            print(f"  ... 另有 {len(preview.to_fill) - 20} 条，见 CSV")

    if preview.to_skip:
        print(f"\n--- 跳过-已一致 ({len(preview.to_skip)}) ---")
        for r in preview.to_skip:
            print(f"  {r.document_number}\t{r.title_en_existing}")

    if preview.conflicts:
        print(f"\n--- 冲突-需人工 ({len(preview.conflicts)}) ---")
        for r in preview.conflicts:
            print(
                f"  {r.document_number}\n"
                f"    台账: {r.title_en_existing}\n"
                f"    外部: {r.title_en_new} [{r.source}]"
            )

    if preview.unmatched:
        print(f"\n--- 外部无台账匹配 ({len(preview.unmatched)}) ---")
        for r in preview.unmatched[:15]:
            print(f"  {r.raw_number}\t{r.title_en}\t[{r.source}]")
        if len(preview.unmatched) > 15:
            print(f"  ... 另有 {len(preview.unmatched) - 15} 条，见 CSV")


def write_csv_reports(preview: PreviewReport, out_dir: Path) -> None:
    out_dir.mkdir(parents=True, exist_ok=True)

    def _write(name: str, headers: list[str], rows: list[tuple]) -> None:
        path = out_dir / name
        with path.open("w", encoding="utf-8-sig", newline="") as f:
            w = csv.writer(f)
            w.writerow(headers)
            w.writerows(rows)

    _write(
        "to_fill.csv",
        ["document_number", "title_zh", "title_en_new", "source", "doc_id"],
        [
            (r.document_number, r.title_zh, r.title_en_new, r.source, r.doc_id)
            for r in preview.to_fill
        ],
    )
    _write(
        "skipped_same.csv",
        ["document_number", "title_zh", "title_en_existing", "source"],
        [
            (r.document_number, r.title_zh, r.title_en_existing, r.source)
            for r in preview.to_skip
        ],
    )
    _write(
        "conflicts.csv",
        ["document_number", "title_zh", "title_en_existing", "title_en_new", "source"],
        [
            (
                r.document_number,
                r.title_zh,
                r.title_en_existing,
                r.title_en_new,
                r.source,
            )
            for r in preview.conflicts
        ],
    )
    _write(
        "unmatched.csv",
        ["raw_number", "title_en", "source"],
        [(r.raw_number, r.title_en, r.source) for r in preview.unmatched],
    )
    print(f"\nCSV 已写入: {out_dir}")


def parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(description="文控台账英文名批量补充")
    p.add_argument("--org-id", default=DEFAULT_ORG_ID, help="目标组织 UUID")
    p.add_argument("--xlsx", default=DEFAULT_XLSX, help="DHF 法规清单 Excel 路径")
    p.add_argument("--docx", default=DEFAULT_DOCX, help="英文目录 Word 路径")
    p.add_argument(
        "--apply",
        action="store_true",
        help="确认写入（默认仅预览）；仅填空 title_en，不覆盖已有",
    )
    p.add_argument(
        "--report-dir",
        default="",
        help="预览 CSV 输出目录（默认 instance/title_en_supplement_<timestamp>）",
    )
    p.add_argument(
        "--no-compare-key",
        action="store_true",
        help="禁用 registration_compare_key 二次匹配",
    )
    return p.parse_args()


def main() -> None:
    args = parse_args()
    xlsx_path = Path(args.xlsx)
    docx_path = Path(args.docx)

    xlsx_map = parse_excel(xlsx_path)
    docx_map = parse_docx(docx_path)
    source_map = merge_sources(xlsx_map, docx_map)
    print(f"[info] Excel 解析 {len(xlsx_map)} 条, Word 解析 {len(docx_map)} 条, 合并 {len(source_map)} 条")

    from webapp import create_app

    app = create_app()
    with app.app_context():
        preview = build_preview(
            args.org_id,
            source_map,
            use_compare_key=not args.no_compare_key,
        )

        if args.report_dir:
            report_dir = Path(args.report_dir)
        else:
            from datetime import datetime

            ts = datetime.now().strftime("%Y%m%d_%H%M%S")
            report_dir = ROOT / "instance" / f"title_en_supplement_{ts}"

        write_csv_reports(preview, report_dir)

        applied: int | None = None
        if args.apply:
            if preview.conflicts:
                print(
                    "\n[WARN] 存在英文冲突记录，--apply 不会覆盖它们，仅填充空白项。",
                    file=sys.stderr,
                )
            applied = apply_fill(args.org_id, preview)
            print_summary(preview, applied=applied)
            if applied:
                print(f"\n[OK] 已更新 {applied} 条 title_en")
            else:
                print("\n[OK] 无需更新（无空白项可填）")
        else:
            print_summary(preview)
            print("\n[dry-run] 未写入数据库。确认后执行: python scripts/supplement_title_en.py --apply")


if __name__ == "__main__":
    main()
